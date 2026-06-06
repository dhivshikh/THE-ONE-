"""
Subject Import Service — Enterprise-grade Excel/CSV bulk import pipeline.

Pipeline:
1. Parse workbook (XLSX/CSV)
2. Validate schema (mandatory columns, correct types)
3. Validate row-level data (uniqueness, FK existence, range checks)
4. Normalize data (map to Subject model fields)
5. Preview import (return row-level results)
6. Commit transaction (batch upsert with rollback safety)
7. Post-import health check (hour coverage, elective integrity)
"""
from __future__ import annotations

import csv
import io
import logging
import traceback
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import and_
from sqlalchemy.orm import Session

logger = logging.getLogger("app.services.subject_import")

# ============================================================================
# COLUMN SPECIFICATION
# ============================================================================

EXPECTED_COLUMNS = [
    "Subject Name",
    "Subject Code",
    "Department",
    "Year",
    "Semester",
    "Theory Hours",
    "Lab Hours",
    "Tutorial Hours",
    "Seminar Hours",
    "Self Study Hours",
    "Project Hours",
    "Report Hours",
    "Is Elective",
    "Assigned Classes",
    "Subject Importance",
    "Previous Year Pass %",
    "Notes / Tags",
]

REQUIRED_COLUMNS = EXPECTED_COLUMNS[:13]  # First 13 are mandatory

COLUMN_ALIAS_MAP = {
    col.lower().strip().replace(" ", "_").replace("/", "_").replace("%", "pct"): col
    for col in EXPECTED_COLUMNS
}

EXPECTED_SHEET_NAME = "SUBJECT_IMPORT"


# ============================================================================
# RESULT DATACLASSES
# ============================================================================

class RowResult:
    """Validation result for a single row."""
    __slots__ = ("row_num", "data", "errors", "warnings", "status", "subject_id")

    def __init__(self, row_num: int, data: dict):
        self.row_num = row_num
        self.data = data
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.status = "pending"  # pending | valid | invalid | imported | updated | skipped
        self.subject_id: Optional[int] = None

    def to_dict(self) -> dict:
        return {
            "row": self.row_num,
            "data": self.data,
            "errors": self.errors,
            "warnings": self.warnings,
            "status": self.status,
            "subject_id": self.subject_id,
        }


class ImportResult:
    """Aggregate import result."""

    def __init__(self):
        self.rows: List[RowResult] = []
        self.schema_errors: List[str] = []
        self.imported = 0
        self.updated = 0
        self.skipped = 0
        self.failed = 0
        self.total_rows = 0
        self.generator_readiness = "unknown"
        self.health_check: dict = {}
        self.batch_id: Optional[str] = None

    @property
    def is_valid(self) -> bool:
        return len(self.schema_errors) == 0 and self.failed == 0

    def to_dict(self) -> dict:
        return {
            "schema_errors": self.schema_errors,
            "total_rows": self.total_rows,
            "imported": self.imported,
            "updated": self.updated,
            "skipped": self.skipped,
            "failed": self.failed,
            "generator_readiness": self.generator_readiness,
            "health_check": self.health_check,
            "batch_id": self.batch_id,
            "rows": [r.to_dict() for r in self.rows],
        }


# ============================================================================
# SERVICE
# ============================================================================

class SubjectImportService:
    """High-performance bulk subject import pipeline."""

    CHUNK_SIZE = 200  # Rows per batch insert

    def __init__(self, db: Session):
        self.db = db
        # Lazy caches — populated once per import
        self._dept_cache: Optional[Dict[str, int]] = None
        self._semester_cache: Optional[Dict[str, Any]] = None  # code -> model
        self._subject_code_cache: Optional[Dict[str, Any]] = None
        self._class_code_cache: Optional[Dict[str, Any]] = None

    # ------------------------------------------------------------------
    # PUBLIC: Full pipeline
    # ------------------------------------------------------------------

    def parse_and_validate(self, file_bytes: bytes, filename: str) -> ImportResult:
        """Parse file, validate schema + rows, return preview (no DB write)."""
        result = ImportResult()
        result.batch_id = f"import_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

        # 1) Parse
        raw_rows = self._parse_file(file_bytes, filename, result)
        if result.schema_errors:
            return result

        # 2) Schema validation (already done in _parse_file)
        # 3) Row-level validation
        self._warm_caches()
        for row in raw_rows:
            self._validate_row(row)

        result.rows = raw_rows
        result.total_rows = len(raw_rows)
        result.failed = sum(1 for r in raw_rows if r.status == "invalid")

        return result

    def commit_import(self, result: ImportResult) -> ImportResult:
        """Commit validated rows to database inside a single transaction."""
        from app.db.models import (
            Subject, SubjectType, Department, Semester,
            subject_semesters, ElectiveBasket,
        )

        if result.schema_errors:
            return result

        valid_rows = [r for r in result.rows if r.status != "invalid"]
        if not valid_rows:
            result.failed = len(result.rows)
            return result

        self._warm_caches()

        try:
            for chunk_start in range(0, len(valid_rows), self.CHUNK_SIZE):
                chunk = valid_rows[chunk_start: chunk_start + self.CHUNK_SIZE]
                for row in chunk:
                    self._upsert_row(row)

            self.db.commit()

            # Count results
            for r in result.rows:
                if r.status == "imported":
                    result.imported += 1
                elif r.status == "updated":
                    result.updated += 1
                elif r.status == "skipped":
                    result.skipped += 1
                elif r.status == "invalid":
                    result.failed += 1

            # Post-import health check
            result.health_check = self._run_health_check()
            result.generator_readiness = (
                "ready" if not result.health_check.get("errors") else "warnings"
            )

        except Exception as e:
            self.db.rollback()
            logger.error(f"Import commit failed — rolled back: {e}\n{traceback.format_exc()}")
            for row in valid_rows:
                if row.status in ("imported", "updated"):
                    row.status = "rollback"
                    row.errors.append(f"Transaction rolled back: {str(e)}")
            result.failed = len(valid_rows)
            result.imported = 0
            result.updated = 0

        return result

    # ------------------------------------------------------------------
    # PARSING
    # ------------------------------------------------------------------

    def _parse_file(self, file_bytes: bytes, filename: str, result: ImportResult) -> List[RowResult]:
        """Parse Excel or CSV file, validate schema, return raw RowResult list."""
        lower = filename.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            return self._parse_excel(file_bytes, result)
        elif lower.endswith(".csv"):
            return self._parse_csv(file_bytes, result)
        else:
            result.schema_errors.append(
                f"Unsupported file format: {filename}. Use .xlsx or .csv"
            )
            return []

    def _parse_excel(self, file_bytes: bytes, result: ImportResult) -> List[RowResult]:
        try:
            import openpyxl
        except ImportError:
            result.schema_errors.append("openpyxl not installed on server")
            return []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            result.schema_errors.append(f"Cannot read Excel file: {e}")
            return []

        # Try to find the expected sheet
        sheet = None
        for name in wb.sheetnames:
            if name.strip().upper() == EXPECTED_SHEET_NAME:
                sheet = wb[name]
                break
        if sheet is None:
            # Fallback: use first sheet
            sheet = wb.active
            if sheet is None:
                result.schema_errors.append("No worksheets found in file")
                return []
            result.rows = []  # reset

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            result.schema_errors.append("Sheet is empty — no header row")
            return []

        headers = [str(h).strip() if h else "" for h in header_row]
        col_map = self._validate_schema(headers, result)
        if result.schema_errors:
            return []

        raw_rows: List[RowResult] = []
        for idx, values in enumerate(rows_iter, start=2):
            # Skip completely empty rows
            if not any(v is not None and str(v).strip() for v in values):
                continue
            data = {}
            for expected_col, col_idx in col_map.items():
                val = values[col_idx] if col_idx < len(values) else None
                data[expected_col] = str(val).strip() if val is not None else ""
            raw_rows.append(RowResult(row_num=idx, data=data))

        return raw_rows

    def _parse_csv(self, file_bytes: bytes, result: ImportResult) -> List[RowResult]:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = file_bytes.decode("latin-1")
            except Exception:
                result.schema_errors.append("Cannot decode CSV file (unsupported encoding)")
                return []

        reader = csv.reader(io.StringIO(text))
        try:
            headers = [h.strip() for h in next(reader)]
        except StopIteration:
            result.schema_errors.append("CSV file is empty — no header row")
            return []

        col_map = self._validate_schema(headers, result)
        if result.schema_errors:
            return []

        raw_rows: List[RowResult] = []
        for idx, values in enumerate(reader, start=2):
            if not any(v.strip() for v in values):
                continue
            data = {}
            for expected_col, col_idx in col_map.items():
                val = values[col_idx] if col_idx < len(values) else ""
                data[expected_col] = val.strip()
            raw_rows.append(RowResult(row_num=idx, data=data))

        return raw_rows

    # ------------------------------------------------------------------
    # SCHEMA VALIDATION
    # ------------------------------------------------------------------

    def _validate_schema(self, headers: List[str], result: ImportResult) -> Dict[str, int]:
        """Validate header row matches expected columns. Return {expected_col: index}."""
        normalized = [h.lower().strip() for h in headers]
        col_map: Dict[str, int] = {}

        for expected_col in EXPECTED_COLUMNS:
            norm_expected = expected_col.lower().strip()
            # Try exact match
            if norm_expected in normalized:
                col_map[expected_col] = normalized.index(norm_expected)
            else:
                # Try partial / fuzzy match
                found = False
                for i, h in enumerate(normalized):
                    # Handle common variations
                    if (norm_expected.replace(" ", "") == h.replace(" ", "")
                            or norm_expected.replace(" ", "_") == h.replace(" ", "_")):
                        col_map[expected_col] = i
                        found = True
                        break
                if not found and expected_col in REQUIRED_COLUMNS:
                    result.schema_errors.append(
                        f"Missing required column: '{expected_col}' (found: {headers})"
                    )
                elif not found:
                    # Optional column — assign sentinel
                    col_map[expected_col] = -1

        return col_map

    # ------------------------------------------------------------------
    # ROW VALIDATION
    # ------------------------------------------------------------------

    def _validate_row(self, row: RowResult):
        """Validate individual row data."""
        d = row.data
        errors = row.errors
        warnings = row.warnings

        # 1) Subject Name required
        if not d.get("Subject Name"):
            errors.append("Subject Name is required")

        # 2) Subject Code required
        code = d.get("Subject Code", "").strip()
        if not code:
            errors.append("Subject Code is required")

        # 3) Department exists (supports semicolon-delimited multi-department)
        dept_str = d.get("Department", "").strip()
        if not dept_str:
            errors.append("Department is required")
        else:
            dept_names = [dn.strip() for dn in dept_str.replace(",", ";").split(";") if dn.strip()]
            # Remove duplicates while preserving order
            seen_dept = set()
            unique_dept_names = []
            for dn in dept_names:
                key = dn.upper()
                if key not in seen_dept:
                    seen_dept.add(key)
                    unique_dept_names.append(dn)
            dept_names = unique_dept_names

            if not dept_names:
                errors.append("Department is required")
            else:
                for dn in dept_names:
                    found = False
                    for key in self._dept_cache:
                        if key.upper() == dn.upper():
                            found = True
                            break
                    if not found:
                        errors.append(f"Department '{dn}' not found in database")

        # 4) Year numeric 1-4
        try:
            year = int(d.get("Year", 0))
            if year < 1 or year > 4:
                errors.append(f"Year must be 1-4, got {year}")
        except (ValueError, TypeError):
            errors.append(f"Year must be numeric, got '{d.get('Year')}'")
            year = 0

        # 5) Semester numeric 1-8
        try:
            semester = int(d.get("Semester", 0))
            if semester < 1 or semester > 8:
                errors.append(f"Semester must be 1-8, got {semester}")
        except (ValueError, TypeError):
            errors.append(f"Semester must be numeric, got '{d.get('Semester')}'")
            semester = 0

        # 6) Hour fields numeric >= 0
        hour_fields = [
            "Theory Hours", "Lab Hours", "Tutorial Hours",
            "Seminar Hours", "Self Study Hours", "Project Hours", "Report Hours"
        ]
        total_hours = 0
        for hf in hour_fields:
            val_str = d.get(hf, "0").strip() or "0"
            try:
                val = int(float(val_str))
                if val < 0:
                    errors.append(f"{hf} must be >= 0, got {val}")
                total_hours += max(0, val)
            except (ValueError, TypeError):
                errors.append(f"{hf} must be numeric, got '{val_str}'")

        # 7) Total hours > 0
        if total_hours == 0 and not errors:
            errors.append("Total weekly hours must be > 0 (at least one component required)")

        # 8) Lab hours divisibility
        lab_str = d.get("Lab Hours", "0").strip() or "0"
        try:
            lab_hours = int(float(lab_str))
            if lab_hours > 0 and lab_hours % 2 != 0:
                errors.append(f"Lab Hours must be even (blocks of 2), got {lab_hours}")
        except (ValueError, TypeError):
            pass

        # 9) Is Elective only Yes/No
        elective_str = d.get("Is Elective", "No").strip().upper()
        if elective_str not in ("YES", "NO", "Y", "N", "TRUE", "FALSE", "1", "0", ""):
            errors.append(f"Is Elective must be Yes/No, got '{d.get('Is Elective')}'")

        # 10) Pass percentage 0-100
        pass_str = d.get("Previous Year Pass %", "").strip()
        if pass_str:
            try:
                pass_pct = int(float(pass_str))
                if pass_pct < 0 or pass_pct > 100:
                    errors.append(f"Pass % must be 0-100, got {pass_pct}")
            except (ValueError, TypeError):
                errors.append(f"Pass % must be numeric, got '{pass_str}'")

        # 11) Subject Importance range
        imp_str = d.get("Subject Importance", "").strip()
        if imp_str:
            try:
                imp = int(float(imp_str))
                if imp < 1 or imp > 5:
                    warnings.append(f"Subject Importance will be clamped to 1-5 range (got {imp})")
            except (ValueError, TypeError):
                # Could be LOW/NORMAL/HIGH text
                if imp_str.upper() not in ("LOW", "NORMAL", "HIGH", ""):
                    warnings.append(f"Subject Importance not recognized: '{imp_str}'")

        # 12) Assigned Classes validation
        classes_str = d.get("Assigned Classes", "").strip()
        if classes_str:
            class_codes = [c.strip() for c in classes_str.replace(",", ";").split(";") if c.strip()]
            # Check for duplicates
            if len(class_codes) != len(set(c.upper() for c in class_codes)):
                errors.append("Duplicate class codes in Assigned Classes")
            # Check existence AND year/semester integrity
            for cc in class_codes:
                if cc.upper() not in {k.upper() for k in self._class_code_cache}:
                    errors.append(f"Class '{cc}' not found in database")
                else:
                    # YEAR/SEMESTER INTEGRITY CHECK
                    sem_obj = next(
                        (v for k, v in self._class_code_cache.items()
                         if k.upper() == cc.upper()), None
                    )
                    if sem_obj and year > 0 and semester > 0:
                        if sem_obj.year != year:
                            errors.append(
                                f"Class '{cc}' is Year {sem_obj.year} but subject is "
                                f"Year {year} — year mismatch. Subject can only be "
                                f"assigned to classes of the same year."
                            )
                        if sem_obj.semester_number != semester:
                            errors.append(
                                f"Class '{cc}' is Semester {sem_obj.semester_number} but "
                                f"subject is Semester {semester} — semester mismatch. "
                                f"Subject can only be assigned to classes of the same semester."
                            )

        # 13) Duplicate code check within current import + DB
        if code and code in self._subject_code_cache:
            existing = self._subject_code_cache[code]
            # Will be treated as update — add warning
            warnings.append(
                f"Subject code '{code}' already exists (ID={existing.id}). "
                f"Will UPDATE if committed."
            )

        row.status = "invalid" if errors else "valid"

    # ------------------------------------------------------------------
    # NORMALIZATION + UPSERT
    # ------------------------------------------------------------------

    def _upsert_row(self, row: RowResult):
        """Normalize and upsert a single row into the database."""
        from app.db.models import (
            Subject, SubjectType, Department, Semester,
            subject_semesters,
        )

        d = row.data
        code = d["Subject Code"].strip()

        # Parse fields
        def safe_int(val, default=0):
            try:
                return int(float(str(val).strip() or default))
            except (ValueError, TypeError):
                return default

        theory = safe_int(d.get("Theory Hours", 0))
        lab = safe_int(d.get("Lab Hours", 0))
        tutorial = safe_int(d.get("Tutorial Hours", 0))
        seminar = safe_int(d.get("Seminar Hours", 0))
        self_study = safe_int(d.get("Self Study Hours", 0))
        project = safe_int(d.get("Project Hours", 0))
        report = safe_int(d.get("Report Hours", 0))
        year = safe_int(d.get("Year", 1), 1)
        semester = safe_int(d.get("Semester", 1), 1)

        elective_str = d.get("Is Elective", "No").strip().upper()
        is_elective = elective_str in ("YES", "Y", "TRUE", "1")

        pass_str = d.get("Previous Year Pass %", "").strip()
        pass_pct = safe_int(pass_str) if pass_str else None

        # Importance mapping
        imp_str = d.get("Subject Importance", "").strip()
        importance = "NORMAL"
        if imp_str:
            try:
                imp_int = int(float(imp_str))
                if imp_int >= 4:
                    importance = "HIGH"
                elif imp_int <= 2:
                    importance = "LOW"
            except (ValueError, TypeError):
                if imp_str.upper() in ("LOW", "NORMAL", "HIGH"):
                    importance = imp_str.upper()

        total_hours = theory + lab + tutorial + seminar + self_study + project + report

        # Resolve departments (supports semicolon-delimited multi-department)
        dept_str = d.get("Department", "").strip()
        dept_ids = []
        dept_objs = []
        if dept_str:
            dept_names = [dn.strip() for dn in dept_str.replace(",", ";").split(";") if dn.strip()]
            # Remove duplicates while preserving order
            seen = set()
            for dn in dept_names:
                if dn.upper() in seen:
                    continue
                seen.add(dn.upper())
                for key, did in self._dept_cache.items():
                    if key.upper() == dn.upper():
                        dept_ids.append(did)
                        break

            # Resolve Department ORM objects for M2M relationship
            if dept_ids:
                from app.db.models import Department as DeptModel
                dept_objs = self.db.query(DeptModel).filter(
                    DeptModel.id.in_(dept_ids)
                ).all()

        # Primary dept_id = first department (backward compat)
        dept_id = dept_ids[0] if dept_ids else None

        # Priority score
        priority_score = Subject.calculate_priority_score(importance, pass_pct)

        # Check if subject exists (for upsert)
        existing = self._subject_code_cache.get(code)

        if existing:
            # UPDATE MODE
            existing.name = d["Subject Name"].strip()
            existing.theory_hours_per_week = theory
            existing.lab_hours_per_week = lab
            existing.tutorial_hours_per_week = tutorial
            existing.seminar_hours_per_week = seminar
            existing.self_study_hours_per_week = self_study
            existing.project_hours_per_week = project
            existing.report_hours_per_week = report
            existing.weekly_hours = total_hours
            existing.year = year
            existing.semester = semester
            existing.is_elective = is_elective
            existing.subject_type = SubjectType.ELECTIVE if is_elective else SubjectType.REGULAR
            existing.dept_id = dept_id or existing.dept_id
            existing.importance_level = importance
            existing.previous_year_pass_percentage = pass_pct
            existing.computed_priority_score = priority_score

            # Update department M2M (replace all)
            if dept_objs:
                existing.departments = dept_objs

            # Update class assignments
            self._assign_classes(existing, d, dept_id)

            row.status = "updated"
            row.subject_id = existing.id
        else:
            # INSERT MODE
            subject = Subject(
                name=d["Subject Name"].strip(),
                code=code,
                theory_hours_per_week=theory,
                lab_hours_per_week=lab,
                tutorial_hours_per_week=tutorial,
                seminar_hours_per_week=seminar,
                self_study_hours_per_week=self_study,
                project_hours_per_week=project,
                report_hours_per_week=report,
                weekly_hours=total_hours,
                year=year,
                semester=semester,
                is_elective=is_elective,
                subject_type=SubjectType.ELECTIVE if is_elective else SubjectType.REGULAR,
                consecutive_slots=1,
                dept_id=dept_id,
                importance_level=importance,
                previous_year_pass_percentage=pass_pct,
                computed_priority_score=priority_score,
            )
            self.db.add(subject)
            self.db.flush()  # Get the ID

            # Assign departments M2M
            if dept_objs:
                subject.departments = dept_objs

            # Assign classes
            self._assign_classes(subject, d, dept_id)

            # Update cache
            self._subject_code_cache[code] = subject

            row.status = "imported"
            row.subject_id = subject.id

    def _assign_classes(self, subject, row_data: dict, dept_id: Optional[int]):
        """Resolve and assign semester (class) records to a subject.
        
        Supports semicolon-delimited class codes: '3A;3B;3C'
        Creates proper M2M entries in subject_semesters for ALL classes.
        
        INTEGRITY RULE: Only assigns classes that match subject's year AND semester.
        Classes with mismatched year/semester are silently skipped (already caught in validation).
        """
        from app.db.models import Semester

        classes_str = row_data.get("Assigned Classes", "").strip()
        if not classes_str:
            return

        class_codes = [c.strip() for c in classes_str.replace(",", ";").split(";") if c.strip()]
        # Remove duplicates while preserving order
        seen = set()
        unique_codes = []
        for cc in class_codes:
            if cc.upper() not in seen:
                seen.add(cc.upper())
                unique_codes.append(cc)
        class_codes = unique_codes

        semester_objs = []
        skipped_mismatches = []
        for cc in class_codes:
            for key, sem_obj in self._class_code_cache.items():
                if key.upper() == cc.upper():
                    # INTEGRITY CHECK: enforce year/semester match at commit time
                    if (subject.year and sem_obj.year != subject.year) or \
                       (subject.semester and sem_obj.semester_number != subject.semester):
                        skipped_mismatches.append(
                            f"{cc} (Year {sem_obj.year}/Sem {sem_obj.semester_number})"
                        )
                    else:
                        semester_objs.append(sem_obj)
                    break

        if skipped_mismatches:
            logger.warning(
                f"Skipped {len(skipped_mismatches)} class(es) for subject '{subject.code}' "
                f"(Year {subject.year}/Sem {subject.semester}) due to year/semester mismatch: "
                f"{skipped_mismatches}"
            )

        if semester_objs:
            # Replace all class assignments (full sync)
            subject.semesters = semester_objs
            self.db.flush()  # Ensure M2M rows are persisted
            logger.info(
                f"Assigned {len(semester_objs)} class(es) to subject '{subject.code}': "
                f"{[s.code for s in semester_objs]}"
            )
            # Auto-assign dept if not set
            if not subject.dept_id and semester_objs:
                subject.dept_id = semester_objs[0].dept_id

    # ------------------------------------------------------------------
    # CACHE WARMING
    # ------------------------------------------------------------------

    def _warm_caches(self):
        """Populate lazy caches for fast lookups."""
        from app.db.models import Department, Semester, Subject

        if self._dept_cache is None:
            depts = self.db.query(Department).all()
            self._dept_cache = {}
            for d in depts:
                self._dept_cache[d.code] = d.id
                self._dept_cache[d.name] = d.id

        if self._subject_code_cache is None:
            subjects = self.db.query(Subject).all()
            self._subject_code_cache = {s.code: s for s in subjects}

        if self._class_code_cache is None:
            semesters = self.db.query(Semester).all()
            self._class_code_cache = {s.code: s for s in semesters}

        if self._semester_cache is None:
            self._semester_cache = self._class_code_cache.copy()

    # ------------------------------------------------------------------
    # HEALTH CHECK
    # ------------------------------------------------------------------

    def _run_health_check(self) -> dict:
        """Run post-import validation checks."""
        from app.db.models import Subject, Semester, subject_semesters

        errors = []
        warnings = []

        # 1) Subjects with zero hours
        zero_hour = self.db.query(Subject).filter(Subject.weekly_hours <= 0).count()
        if zero_hour > 0:
            errors.append(f"{zero_hour} subject(s) have 0 weekly hours")

        # 2) Subjects without class assignments
        orphan_q = (
            self.db.query(Subject)
            .outerjoin(subject_semesters, Subject.id == subject_semesters.c.subject_id)
            .filter(subject_semesters.c.semester_id.is_(None))
            .count()
        )
        if orphan_q > 0:
            warnings.append(f"{orphan_q} subject(s) have no classes assigned")

        # 3) Lab hour sanity
        odd_lab = self.db.query(Subject).filter(
            Subject.lab_hours_per_week > 0,
            Subject.lab_hours_per_week % 2 != 0
        ).count()
        if odd_lab > 0:
            errors.append(f"{odd_lab} subject(s) have odd lab hours (must be even)")

        # 4) Elective subjects without basket
        elective_no_basket = self.db.query(Subject).filter(
            Subject.is_elective == True,
            Subject.elective_basket_id.is_(None)
        ).count()
        if elective_no_basket > 0:
            warnings.append(
                f"{elective_no_basket} elective subject(s) have no basket assigned — "
                f"assign baskets before generation"
            )

        # 5) YEAR/SEMESTER INTEGRITY: Detect cross-year/semester subject-class mappings
        year_sem_mismatches = self.db.execute(
            subject_semesters.select()
        ).fetchall()
        mismatch_count = 0
        mismatch_details = []
        for row in year_sem_mismatches:
            subj = self.db.query(Subject).filter(Subject.id == row.subject_id).first()
            sem = self.db.query(Semester).filter(Semester.id == row.semester_id).first()
            if subj and sem:
                if subj.year != sem.year or subj.semester != sem.semester_number:
                    mismatch_count += 1
                    if len(mismatch_details) < 5:  # Limit detail messages
                        mismatch_details.append(
                            f"{subj.code} (Year {subj.year}/Sem {subj.semester}) → "
                            f"{sem.code} (Year {sem.year}/Sem {sem.semester_number})"
                        )
        if mismatch_count > 0:
            detail_str = "; ".join(mismatch_details)
            if mismatch_count > 5:
                detail_str += f" ... and {mismatch_count - 5} more"
            errors.append(
                f"{mismatch_count} subject-class mapping(s) have year/semester mismatch: "
                f"{detail_str}. Use the integrity repair tool to fix."
            )

        total_subjects = self.db.query(Subject).count()
        total_semesters = self.db.query(Semester).count()

        return {
            "total_subjects": total_subjects,
            "total_classes": total_semesters,
            "errors": errors,
            "warnings": warnings,
            "all_clear": len(errors) == 0,
        }

    # ------------------------------------------------------------------
    # TEMPLATE GENERATION
    # ------------------------------------------------------------------

    @staticmethod
    def generate_template() -> bytes:
        """Generate a downloadable Excel template with SUBJECT_IMPORT sheet."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = EXPECTED_SHEET_NAME

        # Header styling
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_border = Border(
            bottom=Side(style="medium", color="000000"),
            right=Side(style="thin", color="CCCCCC"),
        )
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write headers
        for col_idx, col_name in enumerate(EXPECTED_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border

        # Column widths
        widths = [25, 12, 12, 6, 10, 12, 10, 12, 12, 14, 12, 12, 12, 16, 16, 16, 20]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Example data rows
        examples = [
            ["Data Structures", "CS201", "CSE;IT", 1, 3, 4, 2, 0, 0, 0, 0, 0, "No", "3A;3B", 4, 82, "core"],
            ["Digital Logic Design", "CS202", "CSE", 1, 3, 3, 2, 0, 0, 0, 0, 0, "No", "3A;3B", 5, 76, "important"],
            ["Power BI", "AGI1151", "AIML;AI&DS;IT", 2, 4, 3, 2, 0, 1, 0, 0, 0, "Yes", "4A;4B;4C", 5, 62, "industry"],
        ]

        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row_data in enumerate(examples, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = example_fill
                cell.alignment = Alignment(horizontal="center")

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPECTED_COLUMNS))}1"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
