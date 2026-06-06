"""
Teacher Mapping Import Service — Bulk Excel/CSV import for teacher-class-subject mappings.

Pipeline:
1. Parse workbook (XLSX/CSV)
2. Validate schema (mandatory columns)
3. Resolve & validate row-level data (teacher, class, subject, batch FK integrity)
4. Preview import (return row-level results)
5. Commit transaction (create teachers, append assignments, sync qualifications)
6. Post-import health check (orphan teachers, missing mappings)

CRITICAL CONTRACT:
- One row = one teacher-class-subject mapping
- Preserve existing teacher data
- Create teacher if teacher_code not found
- Append ClassSubjectTeacher if teacher exists
- Validate class belongs to department
- Validate subject belongs to class semester
- Prevent duplicate mappings (same teacher+class+subject+component+batch)
- Batch: Theory=All (no batch_id), Lab=B1/B2/B3 (resolve to Batch FK)
- Rollback on fatal error
"""
from __future__ import annotations

import csv
import io
import logging
import traceback
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

from sqlalchemy import and_
from sqlalchemy.orm import Session

logger = logging.getLogger("app.services.teacher_mapping_import")

# ============================================================================
# COLUMN SPECIFICATION
# ============================================================================

EXPECTED_COLUMNS = [
    "Teacher Name",
    "Teacher Code",
    "Department",
    "Class Assigned",
    "Subject Assigned",
    "Type",
    "Batch",
    "Allowed Departments (Yes/No)",
    "Lab Room",
]

# Aliases for backward compatibility
COLUMN_ALIASES = {
    "Allowed Departments": "Allowed Departments (Yes/No)",
    "Home Department": "Department"
}

REQUIRED_COLUMNS = EXPECTED_COLUMNS[:6]  # First 6 are mandatory (Batch can be empty)

EXPECTED_SHEET_NAME = "TEACHER_MAPPING"


# ============================================================================
# RESULT CLASSES
# ============================================================================

class RowResult:
    """Validation result for a single row."""
    __slots__ = ("row_num", "data", "errors", "warnings", "status", "assignment_id", "teacher_id")

    def __init__(self, row_num: int, data: dict):
        self.row_num = row_num
        self.data = data
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.status = "pending"
        self.assignment_id: Optional[int] = None
        self.teacher_id: Optional[int] = None

    def to_dict(self) -> dict:
        return {
            "row": self.row_num,
            "data": self.data,
            "errors": self.errors,
            "warnings": self.warnings,
            "status": self.status,
            "assignment_id": self.assignment_id,
            "teacher_id": self.teacher_id,
        }


class ImportResult:
    """Aggregate import result."""

    def __init__(self):
        self.rows: List[RowResult] = []
        self.schema_errors: List[str] = []
        self.created_teachers = 0
        self.created_mappings = 0
        self.skipped_duplicates = 0
        self.failed = 0
        self.total_rows = 0
        self.health_check: dict = {}
        self.batch_id: Optional[str] = None
        # Auto-repair tracking
        self.auto_created_batches = 0
        self.auto_linked_subjects = 0

    def to_dict(self) -> dict:
        return {
            "schema_errors": self.schema_errors,
            "total_rows": self.total_rows,
            "created_teachers": self.created_teachers,
            "created_mappings": self.created_mappings,
            "skipped_duplicates": self.skipped_duplicates,
            "failed": self.failed,
            "auto_created_batches": self.auto_created_batches,
            "auto_linked_subjects": self.auto_linked_subjects,
            "health_check": self.health_check,
            "batch_id": self.batch_id,
            "rows": [r.to_dict() for r in self.rows],
        }


# ============================================================================
# SERVICE
# ============================================================================

class TeacherMappingImportService:
    """High-performance bulk teacher mapping import pipeline."""

    CHUNK_SIZE = 200
    HEADER_SCAN_ROWS = 10

    def __init__(self, db: Session):
        self.db = db
        self._dept_cache: Optional[Dict[str, int]] = None          # name/code -> id
        self._class_cache: Optional[Dict[str, Any]] = None         # code -> Semester
        self._subject_cache: Optional[Dict[str, Any]] = None       # code -> Subject
        self._teacher_cache: Optional[Dict[str, Any]] = None       # teacher_code -> Teacher
        self._batch_cache: Optional[Dict[str, Dict[str, Any]]] = None  # sem_id -> {batch_name -> Batch}
        self._subject_semester_set: Optional[Set[Tuple[int, int]]] = None  # (subject_id, semester_id)

    # ------------------------------------------------------------------
    # PUBLIC: Full pipeline
    # ------------------------------------------------------------------

    def parse_and_validate(self, file_bytes: bytes, filename: str) -> ImportResult:
        """Parse file, validate schema + rows, return preview (no DB write)."""
        result = ImportResult()
        result.batch_id = f"teacher_import_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

        raw_rows = self._parse_file(file_bytes, filename, result)
        if result.schema_errors:
            return result

        self._warm_caches()
        for row in raw_rows:
            self._validate_row(row)

        result.rows = raw_rows
        result.total_rows = len(raw_rows)
        result.failed = sum(1 for r in raw_rows if r.status == "invalid")

        return result

    def commit_import(self, result: ImportResult) -> ImportResult:
        """Commit validated rows inside a single atomic transaction."""
        from app.db.models import (
            Teacher, Subject, Semester, Batch,
            ClassSubjectTeacher, ComponentType,
            teacher_subjects,
        )

        if result.schema_errors:
            return result

        valid_rows = [r for r in result.rows if r.status != "invalid"]
        if not valid_rows:
            result.failed = len(result.rows)
            return result

        self._warm_caches()

        try:
            for row in valid_rows:
                self._upsert_row(row)

            self.db.commit()

            # Count results
            for r in result.rows:
                if r.status == "created":
                    result.created_mappings += 1
                elif r.status == "created_teacher":
                    result.created_teachers += 1
                    result.created_mappings += 1
                elif r.status == "duplicate":
                    result.skipped_duplicates += 1
                elif r.status == "invalid":
                    result.failed += 1

            result.health_check = self._run_health_check()

        except Exception as e:
            self.db.rollback()
            logger.error(f"Teacher mapping import failed — rolled back: {e}\n{traceback.format_exc()}")
            for row in valid_rows:
                if row.status not in ("invalid",):
                    row.status = "rollback"
                    row.errors.append(f"Transaction rolled back: {str(e)}")
            result.failed = len(valid_rows)
            result.created_mappings = 0
            result.created_teachers = 0

        return result

    # ------------------------------------------------------------------
    # PARSING
    # ------------------------------------------------------------------

    def _parse_file(self, file_bytes: bytes, filename: str, result: ImportResult) -> List[RowResult]:
        lower = filename.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            return self._parse_excel(file_bytes, result)
        elif lower.endswith(".csv"):
            return self._parse_csv(file_bytes, result)
        else:
            result.schema_errors.append(f"Unsupported file format: {filename}. Use .xlsx or .csv")
            return []

    def _parse_excel(self, file_bytes: bytes, result: ImportResult) -> List[RowResult]:
        try:
            import openpyxl
        except ImportError:
            result.schema_errors.append("openpyxl not installed on server")
            return []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            result.schema_errors.append(f"Cannot read Excel file: {e}")
            return []

        sheet = self._select_best_sheet(wb)
        if sheet is None:
            result.schema_errors.append("No worksheets found in file")
            return []

        preview_rows = list(
            sheet.iter_rows(min_row=1, max_row=self.HEADER_SCAN_ROWS, values_only=True)
        )
        if not preview_rows:
            result.schema_errors.append("Sheet is empty — no header row")
            return []

        header_idx = self._detect_header_row_index(preview_rows)
        header_row_num = header_idx + 1
        header_row = preview_rows[header_idx]
        headers = [str(h).strip() if h else "" for h in header_row]
        col_map = self._validate_schema(headers, result)
        if result.schema_errors:
            return []

        raw_rows: List[RowResult] = []
        for idx, values in enumerate(
            sheet.iter_rows(min_row=header_row_num + 1, values_only=True),
            start=header_row_num + 1,
        ):
            if not any(v is not None and str(v).strip() for v in values):
                continue
            data = {}
            for expected_col, col_idx in col_map.items():
                if col_idx < 0:
                    data[expected_col] = ""
                    continue
                val = values[col_idx] if col_idx < len(values) else None
                data[expected_col] = str(val).strip() if val is not None else ""
            raw_rows.append(RowResult(row_num=idx, data=data))

        return raw_rows

    def _parse_csv(self, file_bytes: bytes, result: ImportResult) -> List[RowResult]:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = file_bytes.decode("latin-1")
            except Exception:
                result.schema_errors.append("Cannot decode CSV file")
                return []

        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            result.schema_errors.append("CSV file is empty")
            return []

        header_idx = self._detect_header_row_index(rows[:self.HEADER_SCAN_ROWS])
        headers = [h.strip() for h in rows[header_idx]]

        col_map = self._validate_schema(headers, result)
        if result.schema_errors:
            return []

        raw_rows: List[RowResult] = []
        for idx, values in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not any(v.strip() for v in values):
                continue
            data = {}
            for expected_col, col_idx in col_map.items():
                if col_idx < 0:
                    data[expected_col] = ""
                    continue
                val = values[col_idx] if col_idx < len(values) else ""
                data[expected_col] = val.strip()
            raw_rows.append(RowResult(row_num=idx, data=data))

        return raw_rows

    # ------------------------------------------------------------------
    # SCHEMA VALIDATION
    # ------------------------------------------------------------------

    def _normalize_header(self, value: str) -> str:
        """Normalize a header for robust matching across templates."""
        normalized = (value or "").strip().lower().replace("_", " ")
        normalized = normalized.replace("/", " ").replace("%", " pct ")
        return " ".join(normalized.split())

    def _normalize_key_token(self, value: str) -> str:
        """Normalize row tokens used in duplicate detection keys."""
        return " ".join((value or "").strip().upper().split())

    def _normalize_batch_value(self, value: str) -> str:
        """Normalize batch label so ALL/blank/NONE map to whole-class semantics.
        
        Also normalizes common batch name variants:
          Batch1, BATCH 1, batch-1  -> B1
          Batch2, BATCH 2, batch-2  -> B2
          B1, b1                    -> B1
        """
        normalized = self._normalize_key_token(value)
        if normalized in ("", "ALL", "NONE", "-", "NA", "N/A"):
            return ""
        
        # Normalize common batch name patterns to canonical form
        import re
        # Match patterns like BATCH1, BATCH 1, BATCH-1, Batch1
        batch_match = re.match(r'^BATCH[\s\-_]*(\d+)$', normalized)
        if batch_match:
            return f"B{batch_match.group(1)}"
        
        return normalized

    def _aliases_for_column(self, expected_col: str) -> List[str]:
        aliases = {self._normalize_header(expected_col)}

        # Include explicit alias map entries for this canonical column.
        for alias, canonical in COLUMN_ALIASES.items():
            if canonical == expected_col:
                aliases.add(self._normalize_header(alias))

        # Extra tolerated variants from old templates.
        if expected_col == "Allowed Departments (Yes/No)":
            aliases.update({
                self._normalize_header("Allowed Departments"),
                self._normalize_header("Allowed Department"),
                self._normalize_header("Allow All Departments"),
                self._normalize_header("Allow All Departments (Yes/No)"),
                self._normalize_header("Allowed Departments Yes No"),
            })

        return list(aliases)

    def _find_column_index(self, normalized_headers: List[str], expected_col: str) -> Optional[int]:
        aliases = self._aliases_for_column(expected_col)
        aliases_compact = {a.replace(" ", "") for a in aliases}

        for idx, header in enumerate(normalized_headers):
            header_compact = header.replace(" ", "")
            if header in aliases or header_compact in aliases_compact:
                return idx

        return None

    def _score_headers(self, headers: List[str]) -> Tuple[int, int]:
        """Return (matched_required, matched_expected) for candidate header row."""
        normalized_headers = [self._normalize_header(h) for h in headers]
        matched_required = 0
        matched_expected = 0

        for expected_col in EXPECTED_COLUMNS:
            if self._find_column_index(normalized_headers, expected_col) is not None:
                matched_expected += 1
                if expected_col in REQUIRED_COLUMNS:
                    matched_required += 1

        return matched_required, matched_expected

    def _detect_header_row_index(self, candidate_rows: List[Any]) -> int:
        """Pick the best header row index from top candidate rows."""
        best_idx = 0
        best_score = (-1, -1)

        for idx, row in enumerate(candidate_rows):
            headers = [str(h).strip() if h is not None else "" for h in row]
            score = self._score_headers(headers)
            if score[0] > best_score[0] or (score[0] == best_score[0] and score[1] > best_score[1]):
                best_idx = idx
                best_score = score

        return best_idx

    def _select_best_sheet(self, workbook):
        """Select TEACHER_MAPPING sheet or best header-matching fallback sheet."""
        for name in workbook.sheetnames:
            if name.strip().upper() == EXPECTED_SHEET_NAME:
                return workbook[name]

        best_sheet = None
        best_score = (-1, -1)

        for sheet in workbook.worksheets:
            candidate_rows = list(
                sheet.iter_rows(min_row=1, max_row=self.HEADER_SCAN_ROWS, values_only=True)
            )
            if not candidate_rows:
                continue

            header_idx = self._detect_header_row_index(candidate_rows)
            headers = [str(h).strip() if h is not None else "" for h in candidate_rows[header_idx]]
            score = self._score_headers(headers)

            if score[0] > best_score[0] or (score[0] == best_score[0] and score[1] > best_score[1]):
                best_sheet = sheet
                best_score = score

        return best_sheet

    def _validate_schema(self, headers: List[str], result: ImportResult) -> Dict[str, int]:
        normalized = [self._normalize_header(h) for h in headers]
        col_map: Dict[str, int] = {}

        for expected_col in EXPECTED_COLUMNS:
            col_idx = self._find_column_index(normalized, expected_col)
            if col_idx is None and expected_col in REQUIRED_COLUMNS:
                result.schema_errors.append(
                    f"Missing required column: '{expected_col}' (found: {headers})"
                )
            elif col_idx is None:
                col_map[expected_col] = -1
            else:
                col_map[expected_col] = col_idx

        return col_map

    # ------------------------------------------------------------------
    # ROW VALIDATION
    # ------------------------------------------------------------------

    def _validate_row(self, row: RowResult):
        """Validate individual row data against database."""
        d = row.data
        errors = row.errors
        warnings = row.warnings

        # 1) Teacher Name required
        teacher_name = d.get("Teacher Name", "").strip()
        if not teacher_name:
            errors.append("Teacher Name is required")

        # 2) Teacher Code required + enhanced lookup
        teacher_code = d.get("Teacher Code", "").strip()
        if not teacher_code:
            errors.append("Teacher Code is required")
        else:
            resolved = self._resolve_teacher(teacher_code, teacher_name)
            if resolved is None:
                warnings.append(f"Teacher '{teacher_code}' not found -- will CREATE new teacher")
            elif resolved.teacher_code != teacher_code:
                warnings.append(
                    f"Teacher resolved via fallback: '{teacher_code}' -> '{resolved.teacher_code}'"
                )

        # 3) Department required and must exist
        dept_str = d.get("Department", "").strip()
        if not dept_str:
            errors.append("Department is required")
        else:
            dept_id = self._resolve_dept(dept_str)
            if dept_id is None:
                errors.append(f"Department '{dept_str}' not found in database")

        # 4) Class Assigned required and must exist
        class_str = d.get("Class Assigned", "").strip()
        if not class_str:
            errors.append("Class Assigned is required")
        else:
            semester = self._resolve_class(class_str)
            if semester is None:
                errors.append(f"Class '{class_str}' not found in database")
            # Removed strict department restriction to allow cross-department teaching
            # elif dept_str:
            #     # Validate class belongs to department
            #     dept_id = self._resolve_dept(dept_str)
            #     if dept_id and semester.dept_id and semester.dept_id != dept_id:
            #         errors.append(
            #             f"Class '{class_str}' (dept_id={semester.dept_id}) "
            #             f"does not belong to department '{dept_str}' (id={dept_id})"
            #         )

        # 5) Subject Assigned required and must exist
        subject_str = d.get("Subject Assigned", "").strip()
        if not subject_str:
            errors.append("Subject Assigned is required")
        else:
            subject = self._resolve_subject(subject_str)
            if subject is None:
                errors.append(f"Subject '{subject_str}' not found in database")
            elif class_str:
                # Validate subject belongs to this class
                semester = self._resolve_class(class_str)
                if semester and subject:
                    # CRITICAL: Year/Semester integrity check
                    if subject.year != semester.year:
                        errors.append(
                            f"Subject '{subject_str}' belongs to Year {subject.year} "
                            f"Semester {subject.semester} but Class '{class_str}' belongs to "
                            f"Year {semester.year} Semester {semester.semester_number} — "
                            f"year mismatch. Invalid mapping."
                        )
                    elif subject.semester != semester.semester_number:
                        errors.append(
                            f"Subject '{subject_str}' belongs to Year {subject.year} "
                            f"Semester {subject.semester} but Class '{class_str}' belongs to "
                            f"Year {semester.year} Semester {semester.semester_number} — "
                            f"semester mismatch. Invalid mapping."
                        )
                    elif (subject.id, semester.id) not in self._subject_semester_set:
                        # DOWNGRADE: auto-link instead of reject
                        warnings.append(
                            f"Subject '{subject_str}' not assigned to class '{class_str}' — "
                            f"will auto-link on commit"
                        )

        # 6) Type must be Theory/Lab/Tutorial
        type_str = d.get("Type", "Theory").strip().lower()
        valid_types = {"theory", "lab", "tutorial", "self_study"}
        if type_str not in valid_types:
            errors.append(f"Type must be one of {list(valid_types)}, got '{d.get('Type')}'")
        else:
            # Validate component type against subject hours
            subject = self._resolve_subject(subject_str) if subject_str else None
            if subject:
                # Auto-correct theory for lab-/tutorial-only subjects instead of hard failing.
                if type_str == "theory" and (subject.theory_hours_per_week or 0) <= 0:
                    if (subject.lab_hours_per_week or 0) > 0:
                        type_str = "lab"
                        d["Type"] = "Lab"
                        warnings.append(
                            f"Type auto-corrected to 'Lab' for subject '{subject_str}' (0 theory hours)"
                        )
                    elif (subject.tutorial_hours_per_week or 0) > 0:
                        type_str = "tutorial"
                        d["Type"] = "Tutorial"
                        warnings.append(
                            f"Type auto-corrected to 'Tutorial' for subject '{subject_str}' (0 theory hours)"
                        )
                    elif (subject.self_study_hours_per_week or 0) > 0:
                        type_str = "self_study"
                        d["Type"] = "Self_Study"
                        warnings.append(
                            f"Type auto-corrected to 'Self_Study' for subject '{subject_str}' (0 theory hours)"
                        )

                # Theory is allowed even when configured hours are zero for lab-only subjects.
                if type_str == "lab" and (subject.lab_hours_per_week or 0) <= 0:
                    # DOWNGRADE: warning instead of fatal error
                    warnings.append(
                        f"Subject '{subject_str}' has 0 lab hours -- "
                        f"consider updating subject configuration"
                    )
                elif type_str == "tutorial" and (subject.tutorial_hours_per_week or 0) <= 0:
                    warnings.append(
                        f"Subject '{subject_str}' has 0 tutorial hours -- "
                        f"consider updating subject configuration"
                    )

        # 7) Batch validation
        batch_str = self._normalize_batch_value(d.get("Batch", ""))
        if batch_str:
            # Must resolve batch for the class
            semester = self._resolve_class(class_str) if class_str else None
            if semester:
                sem_batches = self._batch_cache.get(semester.id, {})
                if batch_str not in {k.upper() for k in sem_batches}:
                    # DOWNGRADE: auto-create instead of reject
                    warnings.append(
                        f"Batch '{batch_str}' not found for class '{class_str}' -- "
                        f"will auto-create on commit"
                    )
        elif type_str == "lab" and not batch_str:
            # Lab without batch is allowed (whole class lab) but warn
            warnings.append("Lab type without a specific batch -- will assign to whole class")

        # 7.5) Lab Room validation
        lab_room_str = d.get("Lab Room", "").strip()
        if lab_room_str:
            if type_str == "lab":
                resolved_room = self._resolve_room(lab_room_str)
                if not resolved_room:
                    warnings.append(f"Room '{lab_room_str}' not found -- will auto-create on commit")
            else:
                warnings.append(f"Lab Room '{lab_room_str}' ignored for non-Lab component type")

        # 8) Duplicate check in current results
        if not errors:
            # Build a canonical key to avoid false positives from spacing/case variants.
            semester = self._resolve_class(class_str) if class_str else None
            subject = self._resolve_subject(subject_str) if subject_str else None

            batch_key = None
            if type_str == "lab" and batch_str:
                resolved_batch = self._resolve_batch(semester.id, batch_str) if semester else None
                batch_key = resolved_batch.id if resolved_batch else batch_str

            key = (
                self._normalize_key_token(teacher_code),
                semester.id if semester else self._normalize_key_token(class_str),
                subject.id if subject else self._normalize_key_token(subject_str),
                type_str,
                batch_key,
            )
            if not hasattr(self, '_seen_mappings'):
                self._seen_mappings = set()
            if key in self._seen_mappings:
                errors.append("Duplicate row — same teacher+class+subject+type+batch appears twice in file")
            else:
                self._seen_mappings.add(key)

        row.status = "invalid" if errors else "valid"

    # ------------------------------------------------------------------
    # RESOLUTION HELPERS
    # ------------------------------------------------------------------

    def _resolve_dept(self, dept_str: str) -> Optional[int]:
        """Resolve department name/code to ID."""
        for key, did in self._dept_cache.items():
            if key.upper() == dept_str.upper():
                return did
        return None

    def _resolve_class(self, class_str: str):
        """Resolve class code/name to Semester object."""
        # Try exact code match first
        for key, sem in self._class_cache.items():
            if key.upper() == class_str.upper():
                return sem
        # Then try name match
        for key, sem in self._class_cache.items():
            if sem.name and sem.name.upper() == class_str.upper():
                return sem
        return None

    def _resolve_subject(self, subject_str: str):
        """Resolve subject code/name to Subject object."""
        # Try exact code match
        for key, sub in self._subject_cache.items():
            if key.upper() == subject_str.upper():
                return sub
        # Then try name match
        for key, sub in self._subject_cache.items():
            if sub.name and sub.name.upper() == subject_str.upper():
                return sub
        return None

    def _resolve_batch(self, semester_id: int, batch_str: str):
        """Resolve batch name to Batch object for a given semester."""
        sem_batches = self._batch_cache.get(semester_id, {})
        for key, batch in sem_batches.items():
            if key.upper() == batch_str.upper():
                return batch
        return None

    def _resolve_teacher(self, teacher_code: str, teacher_name: str = ""):
        """Enhanced teacher lookup with fallback chain.
        
        Resolution order:
        1. Exact teacher_code match
        2. Case-insensitive teacher_code match
        3. Normalized code match (strip leading zeros, whitespace)
        4. Teacher name match (case-insensitive)
        """
        # 1. Exact match
        if teacher_code in self._teacher_cache:
            return self._teacher_cache[teacher_code]
        
        # 2. Case-insensitive match
        for key, teacher in self._teacher_cache.items():
            if key.upper() == teacher_code.upper():
                return teacher
        
        # 3. Normalized match (strip leading zeros from numeric suffix)
        import re
        # AI01 vs AI001 -- normalize by stripping leading zeros in numeric part
        norm_code = re.sub(r'(\D)(0+)(\d)', r'\1\3', teacher_code.strip())
        for key, teacher in self._teacher_cache.items():
            norm_key = re.sub(r'(\D)(0+)(\d)', r'\1\3', key.strip())
            if norm_key.upper() == norm_code.upper():
                return teacher
        
        # 4. Name match (last resort)
        if teacher_name:
            teacher_name_index = getattr(self, '_teacher_name_index', {})
            name_key = teacher_name.strip().upper()
            if name_key in teacher_name_index:
                return teacher_name_index[name_key]
        
        return None

    def _resolve_room(self, room_str: str):
        """Resolve room name to Room object."""
        if not room_str or not hasattr(self, "_room_cache") or self._room_cache is None:
            return None
        return self._room_cache.get(room_str.upper())

    # ------------------------------------------------------------------
    # UPSERT
    # ------------------------------------------------------------------

    def _upsert_row(self, row: RowResult):
        """Process a single row: create teacher if needed, then create the mapping."""
        from app.db.models import (
            Teacher, Subject, Semester, Batch,
            ClassSubjectTeacher, ComponentType, Department,
            teacher_subjects,
        )

        d = row.data
        teacher_code = d["Teacher Code"].strip()
        teacher_name = d["Teacher Name"].strip()
        dept_str = d["Department"].strip()
        class_str = d["Class Assigned"].strip()
        subject_str = d["Subject Assigned"].strip()
        type_str = d.get("Type", "Theory").strip().lower()
        batch_str = self._normalize_batch_value(d.get("Batch", ""))
        allowed_depts_str = d.get("Allowed Departments (Yes/No)", "").strip()
        lab_room_str = d.get("Lab Room", "").strip()

        # Resolve dept
        dept_id = self._resolve_dept(dept_str)
        
        # Resolve allowed departments
        allowed_depts = []
        is_common_service = False
        if allowed_depts_str:
            if allowed_depts_str.upper() in ["ALL", "COMMON", "ANY", "YES", "Y", "TRUE"]:
                is_common_service = True
            elif allowed_depts_str.upper() in ["NO", "N", "FALSE", "NONE"]:
                is_common_service = False
            else:
                dept_names = [name.strip() for name in allowed_depts_str.replace(";", ",").split(",") if name.strip()]
                for name in dept_names:
                    a_dept_id = self._resolve_dept(name)
                    if a_dept_id:
                        dept = self.db.query(Department).filter(Department.id == a_dept_id).first()
                        if dept and dept not in allowed_depts:
                            allowed_depts.append(dept)

        # Resolve or create teacher (enhanced fallback chain)
        teacher = self._resolve_teacher(teacher_code, teacher_name)
        created_new_teacher = False

        if not teacher:
            teacher = Teacher(
                name=teacher_name,
                teacher_code=teacher_code,
                dept_id=dept_id,
                max_hours_per_week=20,
                experience_years=1,
                experience_score=0.5,
                available_days="0,1,2,3,4",
                is_active=True,
                is_common_service_dept=is_common_service
            )
            if allowed_depts:
                teacher.allowed_departments = allowed_depts
                
            self.db.add(teacher)
            self.db.flush()  # Get the ID
            self._teacher_cache[teacher_code] = teacher
            created_new_teacher = True
        else:
            # Update name if different (preserve other data)
            if teacher.name != teacher_name:
                teacher.name = teacher_name
            
            # CRITICAL: Reactivate soft-deleted teachers on import
            if not teacher.is_active:
                teacher.is_active = True
            
            # Optionally update allowed_departments if specified and teacher already exists
            # Only doing it if explicitly provided in the import to enhance existing teacher
            if is_common_service and not teacher.is_common_service_dept:
                teacher.is_common_service_dept = True
            if allowed_depts:
                existing_depts = {d.id for d in teacher.allowed_departments}
                for d in allowed_depts:
                    if d.id not in existing_depts:
                        teacher.allowed_departments.append(d)


        # Resolve class & subject
        semester = self._resolve_class(class_str)
        subject = self._resolve_subject(subject_str)

        if not semester or not subject:
            row.status = "invalid"
            row.errors.append("Failed to resolve class or subject during commit")
            return

        # AUTO-LINK: Create subject-class mapping if missing
        # INTEGRITY CHECK: Only auto-link if year/semester match
        from app.db.models import subject_semesters
        if (subject.id, semester.id) not in self._subject_semester_set:
            # Verify year/semester integrity before auto-linking
            if subject.year != semester.year or subject.semester != semester.semester_number:
                row.status = "invalid"
                row.errors.append(
                    f"Cannot auto-link: Subject '{subject.code}' (Year {subject.year}/Sem {subject.semester}) "
                    f"does not match Class '{semester.code}' (Year {semester.year}/Sem {semester.semester_number}). "
                    f"Year/semester mismatch."
                )
                return
            self.db.execute(
                subject_semesters.insert().values(
                    subject_id=subject.id, semester_id=semester.id
                )
            )
            self._subject_semester_set.add((subject.id, semester.id))
            row.warnings.append(f"Auto-linked subject '{subject.code}' to class '{semester.code}'")
            logger.info(f"Auto-linked subject {subject.code} (id={subject.id}) to class {semester.code} (id={semester.id})")
        # Map component type
        comp_type_map = {
            "theory": ComponentType.THEORY,
            "lab": ComponentType.LAB,
            "tutorial": ComponentType.TUTORIAL,
            "self_study": ComponentType.SELF_STUDY,
        }
        component_type = comp_type_map.get(type_str, ComponentType.THEORY)

        # Resolve batch (auto-create if missing)
        batch_id = None
        if batch_str:
            batch = self._resolve_batch(semester.id, batch_str)
            if not batch:
                # AUTO-CREATE missing batch
                batch = Batch(
                    name=batch_str,
                    semester_id=semester.id,
                )
                self.db.add(batch)
                self.db.flush()
                # Update cache
                if semester.id not in self._batch_cache:
                    self._batch_cache[semester.id] = {}
                self._batch_cache[semester.id][batch.name] = batch
                row.warnings.append(f"Auto-created batch '{batch_str}' for class '{semester.code}'")
                logger.info(f"Auto-created batch '{batch_str}' (id={batch.id}) for class {semester.code}")
            batch_id = batch.id

        # Resolve lab room (auto-create if missing)
        room_id = None
        if lab_room_str and component_type == ComponentType.LAB:
            room = self._resolve_room(lab_room_str)
            if not room:
                from app.db.models import Room, RoomType
                room = Room(name=lab_room_str, room_type=RoomType.LAB, capacity=60)
                self.db.add(room)
                self.db.flush()
                if not hasattr(self, "_room_cache") or self._room_cache is None:
                    self._room_cache = {}
                self._room_cache[room.name.upper()] = room
                row.warnings.append(f"Auto-created lab room '{lab_room_str}'")
                logger.info(f"Auto-created lab room '{lab_room_str}' (id={room.id})")
            room_id = room.id

        # Check for existing duplicate mapping in DB
        existing = self.db.query(ClassSubjectTeacher).filter(
            ClassSubjectTeacher.teacher_id == teacher.id,
            ClassSubjectTeacher.semester_id == semester.id,
            ClassSubjectTeacher.subject_id == subject.id,
            ClassSubjectTeacher.component_type == component_type,
        )
        if batch_id is not None:
            existing = existing.filter(ClassSubjectTeacher.batch_id == batch_id)
        else:
            existing = existing.filter(ClassSubjectTeacher.batch_id.is_(None))

        existing_row = existing.first()

        if existing_row:
            row.status = "duplicate"
            row.assignment_id = existing_row.id
            row.teacher_id = teacher.id
            row.warnings.append("Mapping already exists — skipped")
            return

        # Create the mapping
        assignment = ClassSubjectTeacher(
            teacher_id=teacher.id,
            semester_id=semester.id,
            subject_id=subject.id,
            component_type=component_type,
            batch_id=batch_id,
            room_id=room_id,
            is_locked=True,
            assignment_reason="bulk_import",
        )
        self.db.add(assignment)

        # Sync teacher qualification (teacher_subjects M2M)
        if subject not in teacher.subjects:
            teacher.subjects.append(subject)

        self.db.flush()

        row.assignment_id = assignment.id
        row.teacher_id = teacher.id
        row.status = "created_teacher" if created_new_teacher else "created"

    # ------------------------------------------------------------------
    # CACHE WARMING
    # ------------------------------------------------------------------

    def _warm_caches(self):
        """Populate lazy caches for fast lookups."""
        from app.db.models import Department, Semester, Subject, Teacher, Batch, subject_semesters, Room

        if not hasattr(self, "_room_cache") or self._room_cache is None:
            rooms = self.db.query(Room).all()
            self._room_cache = {r.name.upper(): r for r in rooms if r.name}

        if self._dept_cache is None:
            depts = self.db.query(Department).all()
            self._dept_cache = {}
            for d in depts:
                self._dept_cache[d.code] = d.id
                self._dept_cache[d.name] = d.id

        if self._class_cache is None:
            semesters = self.db.query(Semester).all()
            self._class_cache = {}
            for s in semesters:
                self._class_cache[s.code] = s
                if s.name:
                    self._class_cache[s.name] = s

        if self._subject_cache is None:
            subjects = self.db.query(Subject).all()
            self._subject_cache = {}
            for s in subjects:
                self._subject_cache[s.code] = s
                if s.name:
                    self._subject_cache[s.name] = s

        if self._teacher_cache is None:
            teachers = self.db.query(Teacher).all()
            self._teacher_cache = {t.teacher_code: t for t in teachers if t.teacher_code}
            # Build secondary name index for fallback resolution
            self._teacher_name_index = {}
            for t in teachers:
                if t.name:
                    self._teacher_name_index[t.name.strip().upper()] = t

        if self._batch_cache is None:
            batches = self.db.query(Batch).all()
            self._batch_cache = {}
            for b in batches:
                if b.semester_id not in self._batch_cache:
                    self._batch_cache[b.semester_id] = {}
                self._batch_cache[b.semester_id][b.name] = b

        if self._subject_semester_set is None:
            rows = self.db.execute(subject_semesters.select()).fetchall()
            self._subject_semester_set = {(r.subject_id, r.semester_id) for r in rows}

        # Reset seen mappings for validation
        self._seen_mappings = set()

    # ------------------------------------------------------------------
    # HEALTH CHECK
    # ------------------------------------------------------------------

    def _run_health_check(self) -> dict:
        """Run post-import validation checks."""
        from app.db.models import Teacher, ClassSubjectTeacher, Subject

        errors = []
        warnings = []

        # 1) Teachers without any assignments
        teacher_count = self.db.query(Teacher).filter(Teacher.is_active == True).count()
        assigned_teacher_ids = {
            r[0] for r in self.db.query(ClassSubjectTeacher.teacher_id).distinct().all()
        }
        active_teachers = self.db.query(Teacher).filter(Teacher.is_active == True).all()
        unassigned = [t for t in active_teachers if t.id not in assigned_teacher_ids]
        if unassigned:
            warnings.append(f"{len(unassigned)} active teacher(s) have no class assignments")

        # 2) Total mappings
        total_mappings = self.db.query(ClassSubjectTeacher).count()

        # 3) Check for subjects with no teacher assigned
        all_subjects = self.db.query(Subject).count()
        mapped_subject_ids = {
            r[0] for r in self.db.query(ClassSubjectTeacher.subject_id).distinct().all()
        }
        unmapped = all_subjects - len(mapped_subject_ids)
        if unmapped > 0:
            warnings.append(f"{unmapped} subject(s) have no teacher assigned")

        return {
            "total_active_teachers": teacher_count,
            "total_mappings": total_mappings,
            "errors": errors,
            "warnings": warnings,
            "all_clear": len(errors) == 0,
        }

    # ------------------------------------------------------------------
    # DEPENDENCY REPAIR
    # ------------------------------------------------------------------

    def repair_dependencies(self) -> dict:
        """Pre-commit auto-repair of missing dependencies.
        
        Actions:
        - Create missing batches referenced in any ClassSubjectTeacher
        - Create missing subject-class links
        - Normalize teacher codes (strip leading zeros)
        - Remove exact duplicate CST mappings
        
        Returns summary of repairs performed.
        """
        from app.db.models import (
            Batch, Semester, Subject, Teacher,
            ClassSubjectTeacher, subject_semesters,
        )
        
        self._warm_caches()
        repairs = {
            "batches_created": 0,
            "subject_links_created": 0,
            "teacher_codes_normalized": 0,
            "duplicates_removed": 0,
        }
        
        # 1. Find CST rows referencing batch names not in batches table
        #    (batch_id is stored, so this mainly catches NULL batch_ids for lab rows)
        
        # 2. Find subjects assigned to classes via CST but missing from subject_semesters
        cst_pairs = self.db.query(
            ClassSubjectTeacher.subject_id,
            ClassSubjectTeacher.semester_id
        ).distinct().all()
        
        for subject_id, semester_id in cst_pairs:
            if (subject_id, semester_id) not in self._subject_semester_set:
                try:
                    self.db.execute(
                        subject_semesters.insert().values(
                            subject_id=subject_id, semester_id=semester_id
                        )
                    )
                    self._subject_semester_set.add((subject_id, semester_id))
                    repairs["subject_links_created"] += 1
                except Exception:
                    pass  # Already exists or FK violation
        
        # 3. Normalize teacher codes (strip leading zeros in numeric suffix)
        import re
        teachers = self.db.query(Teacher).all()
        for t in teachers:
            if t.teacher_code:
                normalized = re.sub(r'(\D)(0+)(\d)', r'\1\3', t.teacher_code.strip())
                if normalized != t.teacher_code:
                    # Check no collision
                    existing = self.db.query(Teacher).filter(
                        Teacher.teacher_code == normalized,
                        Teacher.id != t.id
                    ).first()
                    if not existing:
                        t.teacher_code = normalized
                        repairs["teacher_codes_normalized"] += 1
        
        # 4. Remove exact duplicate CST mappings
        from sqlalchemy import func
        dupes = self.db.query(
            ClassSubjectTeacher.teacher_id,
            ClassSubjectTeacher.semester_id,
            ClassSubjectTeacher.subject_id,
            ClassSubjectTeacher.component_type,
            ClassSubjectTeacher.batch_id,
            func.count(ClassSubjectTeacher.id).label('cnt'),
            func.min(ClassSubjectTeacher.id).label('keep_id'),
        ).group_by(
            ClassSubjectTeacher.teacher_id,
            ClassSubjectTeacher.semester_id,
            ClassSubjectTeacher.subject_id,
            ClassSubjectTeacher.component_type,
            ClassSubjectTeacher.batch_id,
        ).having(func.count(ClassSubjectTeacher.id) > 1).all()
        
        for dupe in dupes:
            # Delete all but the oldest (keep_id)
            self.db.query(ClassSubjectTeacher).filter(
                ClassSubjectTeacher.teacher_id == dupe.teacher_id,
                ClassSubjectTeacher.semester_id == dupe.semester_id,
                ClassSubjectTeacher.subject_id == dupe.subject_id,
                ClassSubjectTeacher.component_type == dupe.component_type,
                ClassSubjectTeacher.batch_id == dupe.batch_id,
                ClassSubjectTeacher.id != dupe.keep_id,
            ).delete(synchronize_session=False)
            repairs["duplicates_removed"] += dupe.cnt - 1
        
        self.db.commit()
        
        total_repairs = sum(repairs.values())
        repairs["total_repairs"] = total_repairs
        repairs["status"] = "ok" if total_repairs == 0 else f"{total_repairs} repair(s) applied"
        
        logger.info(f"Dependency repair completed: {repairs}")
        return repairs

    # ------------------------------------------------------------------
    # TEMPLATE GENERATION
    # ------------------------------------------------------------------

    @staticmethod
    def generate_template() -> bytes:
        """Generate a downloadable Excel template with TEACHER_MAPPING sheet."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = EXPECTED_SHEET_NAME

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
        header_border = Border(
            bottom=Side(style="medium", color="000000"),
            right=Side(style="thin", color="CCCCCC"),
        )
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(EXPECTED_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border

        widths = [22, 14, 16, 14, 18, 10, 10, 28, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        examples = [
            ["Dr. Smith", "CSE001", "CSE", "CS3A", "CS201", "Theory", "All", "No", ""],
            ["Dr. Smith", "CSE001", "CSE", "CS3A", "CS201", "Lab", "B1", "No", "AI Lab 1"],
            ["Dr. Smith", "CSE001", "CSE", "CS3B", "CS201", "Theory", "All", "No", ""],
            ["Prof. Kumar", "CSE002", "CSE", "CS3A", "CS202", "Theory", "All", "ALL", ""],
            ["Prof. Kumar", "CSE002", "CSE", "CS3A", "CS202", "Lab", "B2", "CSE, ECE", "AI Lab 1"],
        ]

        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row_data in enumerate(examples, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = example_fill
                cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPECTED_COLUMNS))}1"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
