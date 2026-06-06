import io
import csv
import logging
import traceback
from datetime import datetime
from typing import List, Dict, Optional, Any

from sqlalchemy.orm import Session

from app.db.models import Department
from app.services.subject_import_service import RowResult, ImportResult

logger = logging.getLogger("app.services.department_import")

EXPECTED_COLUMNS = [
    "Department Name",
    "Department Code",
    "HOD",
    "Years Available",
    "Semesters Available",
    "Notes"
]

REQUIRED_COLUMNS = ["Department Name", "Department Code"]

EXPECTED_SHEET_NAME = "DEPARTMENTS"


class DepartmentImportService:
    CHUNK_SIZE = 200

    def __init__(self, db: Session):
        self.db = db
        self._dept_code_cache: Optional[Dict[str, Department]] = None

    def parse_and_validate(self, file_bytes: bytes, filename: str) -> ImportResult:
        result = ImportResult()
        result.batch_id = f"dept_import_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

        raw_rows = self._parse_file(file_bytes, filename, result)
        if result.schema_errors:
            return result

        self._warm_caches()
        for row in raw_rows:
            self._validate_row(row)

        result.rows = raw_rows
        result.total_rows = len(raw_rows)
        result.failed = sum(1 for r in raw_rows if r.status == "invalid")
        return result

    def commit_import(self, result: ImportResult) -> ImportResult:
        if result.schema_errors:
            return result

        valid_rows = [r for r in result.rows if r.status != "invalid"]
        if not valid_rows:
            result.failed = len(result.rows)
            return result

        self._warm_caches()

        try:
            for chunk_start in range(0, len(valid_rows), self.CHUNK_SIZE):
                chunk = valid_rows[chunk_start: chunk_start + self.CHUNK_SIZE]
                for row in chunk:
                    self._upsert_row(row)

            self.db.commit()

            for r in result.rows:
                if r.status == "imported":
                    result.imported += 1
                elif r.status == "updated":
                    result.updated += 1
                elif r.status == "invalid":
                    result.failed += 1

            result.health_check = self._run_health_check()
            result.generator_readiness = "ready" if not result.health_check.get("errors") else "warnings"
        except Exception as e:
            self.db.rollback()
            logger.error(f"Import commit failed: {e}\n{traceback.format_exc()}")
            for row in valid_rows:
                if row.status in ("imported", "updated"):
                    row.status = "rollback"
                    row.errors.append(f"Transaction rolled back: {str(e)}")
            result.failed = len(valid_rows)
            result.imported = 0
            result.updated = 0

        return result

    def _parse_file(self, file_bytes: bytes, filename: str, result: ImportResult) -> List[RowResult]:
        lower = filename.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            return self._parse_excel(file_bytes, result)
        elif lower.endswith(".csv"):
            return self._parse_csv(file_bytes, result)
        else:
            result.schema_errors.append("Unsupported file format. Use .xlsx or .csv")
            return []

    def _parse_excel(self, file_bytes: bytes, result: ImportResult) -> List[RowResult]:
        try:
            import openpyxl
        except ImportError:
            result.schema_errors.append("openpyxl not installed on server")
            return []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            result.schema_errors.append(f"Cannot read Excel file: {e}")
            return []

        sheet = None
        for name in wb.sheetnames:
            if name.strip().upper() == EXPECTED_SHEET_NAME:
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active
            if sheet is None:
                result.schema_errors.append("No worksheets found in file")
                return []

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            result.schema_errors.append("Sheet is empty — no header row")
            return []

        headers = [str(h).strip() if h else "" for h in header_row]
        col_map = self._validate_schema(headers, result)
        if result.schema_errors:
            return []

        raw_rows: List[RowResult] = []
        for idx, values in enumerate(rows_iter, start=2):
            if not any(v is not None and str(v).strip() for v in values):
                continue
            data = {}
            for expected_col, col_idx in col_map.items():
                if col_idx == -1:
                    data[expected_col] = ""
                    continue
                val = values[col_idx] if col_idx < len(values) else None
                data[expected_col] = str(val).strip() if val is not None else ""
            raw_rows.append(RowResult(row_num=idx, data=data))

        return raw_rows

    def _parse_csv(self, file_bytes: bytes, result: ImportResult) -> List[RowResult]:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = file_bytes.decode("latin-1")
            except Exception:
                result.schema_errors.append("Cannot decode CSV file (unsupported encoding)")
                return []

        reader = csv.reader(io.StringIO(text))
        try:
            headers = [h.strip() for h in next(reader)]
        except StopIteration:
            result.schema_errors.append("CSV file is empty — no header row")
            return []

        col_map = self._validate_schema(headers, result)
        if result.schema_errors:
            return []

        raw_rows: List[RowResult] = []
        for idx, values in enumerate(reader, start=2):
            if not any(v.strip() for v in values):
                continue
            data = {}
            for expected_col, col_idx in col_map.items():
                if col_idx == -1:
                    data[expected_col] = ""
                    continue
                val = values[col_idx] if col_idx < len(values) else ""
                data[expected_col] = val.strip()
            raw_rows.append(RowResult(row_num=idx, data=data))

        return raw_rows

    def _validate_schema(self, headers: List[str], result: ImportResult) -> Dict[str, int]:
        normalized = [h.lower().strip() for h in headers]
        col_map: Dict[str, int] = {}
        
        for expected_col in EXPECTED_COLUMNS:
            norm_expected = expected_col.lower().strip()
            if norm_expected in normalized:
                col_map[expected_col] = normalized.index(norm_expected)
            else:
                found = False
                for i, h in enumerate(normalized):
                    if norm_expected.replace(" ", "") == h.replace(" ", "") or norm_expected.replace(" ", "_") == h.replace(" ", "_"):
                        col_map[expected_col] = i
                        found = True
                        break
                if not found and expected_col in REQUIRED_COLUMNS:
                    result.schema_errors.append(f"Missing required column: '{expected_col}' (found: {headers})")
                elif not found:
                    col_map[expected_col] = -1

        return col_map

    def _validate_row(self, row: RowResult):
        d = row.data
        errors = row.errors
        warnings = row.warnings

        name = d.get("Department Name", "").strip()
        if not name:
            errors.append("Department Name is required")

        code = d.get("Department Code", "").strip()
        if not code:
            errors.append("Department Code is required")

        if code and code.upper() in self._dept_code_cache:
            warnings.append(f"Department code '{code}' already exists. Will UPDATE if committed.")

        row.status = "invalid" if errors else "valid"

    def _upsert_row(self, row: RowResult):
        d = row.data
        code = d["Department Code"].strip()
        name = d["Department Name"].strip()

        existing = self._dept_code_cache.get(code.upper())
        if existing:
            existing.name = name
            row.status = "updated"
        else:
            dept = Department(name=name, code=code)
            self.db.add(dept)
            self.db.flush()
            self._dept_code_cache[code.upper()] = dept
            row.status = "imported"

    def _warm_caches(self):
        if self._dept_code_cache is None:
            depts = self.db.query(Department).all()
            self._dept_code_cache = {d.code.upper(): d for d in depts}

    def _run_health_check(self) -> dict:
        total = self.db.query(Department).count()
        return {
            "total_departments": total,
            "errors": [],
            "warnings": [],
            "all_clear": True,
        }

    @staticmethod
    def generate_template() -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = EXPECTED_SHEET_NAME

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(EXPECTED_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for i, w in enumerate([25, 15, 20, 15, 20, 25], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        examples = [
            ["Computer Science", "CSE", "Dr Kumar", "1,2,3,4", "1,3,5,7", "Core department"],
            ["AI & ML", "AIML", "Dr Rao", "1,2,3,4", "1,3,5,7", "Newer branch"],
        ]
        
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row_data in enumerate(examples, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = example_fill
                cell.alignment = Alignment(horizontal="center")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPECTED_COLUMNS))}1"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
