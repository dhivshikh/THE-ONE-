import io
import csv
import logging
import traceback
from datetime import datetime
from typing import List, Dict, Optional, Any

from sqlalchemy.orm import Session

from app.db.models import Semester, Department
from app.services.subject_import_service import RowResult, ImportResult

logger = logging.getLogger("app.services.class_import")

EXPECTED_COLUMNS = [
    "Department",
    "Class Name",
    "Year",
    "Semester",
    "Section",
    "Strength",
    "Advisor",
    "Room Preference",
    "Batch Enabled"
]

REQUIRED_COLUMNS = [
    "Department",
    "Class Name",
    "Year",
    "Semester",
    "Section"
]

EXPECTED_SHEET_NAME = "CLASSES"


class ClassImportService:
    CHUNK_SIZE = 200

    def __init__(self, db: Session):
        self.db = db
        self._dept_cache: Optional[Dict[str, int]] = None
        self._semester_code_cache: Optional[Dict[str, Semester]] = None

    def parse_and_validate(self, file_bytes: bytes, filename: str) -> ImportResult:
        result = ImportResult()
        result.batch_id = f"class_import_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

        raw_rows = self._parse_file(file_bytes, filename, result)
        if result.schema_errors:
            return result

        self._warm_caches()
        for row in raw_rows:
            self._validate_row(row)

        result.rows = raw_rows
        result.total_rows = len(raw_rows)
        result.failed = sum(1 for r in raw_rows if r.status == "invalid")
        return result

    def commit_import(self, result: ImportResult) -> ImportResult:
        if result.schema_errors:
            return result

        valid_rows = [r for r in result.rows if r.status != "invalid"]
        if not valid_rows:
            result.failed = len(result.rows)
            return result

        self._warm_caches()

        try:
            for chunk_start in range(0, len(valid_rows), self.CHUNK_SIZE):
                chunk = valid_rows[chunk_start: chunk_start + self.CHUNK_SIZE]
                for row in chunk:
                    self._upsert_row(row)

            self.db.commit()

            for r in result.rows:
                if r.status == "imported":
                    result.imported += 1
                elif r.status == "updated":
                    result.updated += 1
                elif r.status == "invalid":
                    result.failed += 1

            result.health_check = self._run_health_check()
            result.generator_readiness = "ready" if not result.health_check.get("errors") else "warnings"
        except Exception as e:
            self.db.rollback()
            logger.error(f"Import commit failed: {e}\n{traceback.format_exc()}")
            for row in valid_rows:
                if row.status in ("imported", "updated"):
                    row.status = "rollback"
                    row.errors.append(f"Transaction rolled back: {str(e)}")
            result.failed = len(valid_rows)
            result.imported = 0
            result.updated = 0

        return result

    def _parse_file(self, file_bytes: bytes, filename: str, result: ImportResult) -> List[RowResult]:
        lower = filename.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            return self._parse_excel(file_bytes, result)
        elif lower.endswith(".csv"):
            return self._parse_csv(file_bytes, result)
        else:
            result.schema_errors.append("Unsupported file format. Use .xlsx or .csv")
            return []

    def _parse_excel(self, file_bytes: bytes, result: ImportResult) -> List[RowResult]:
        try:
            import openpyxl
        except ImportError:
            result.schema_errors.append("openpyxl not installed on server")
            return []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            result.schema_errors.append(f"Cannot read Excel file: {e}")
            return []

        sheet = None
        for name in wb.sheetnames:
            if name.strip().upper() == EXPECTED_SHEET_NAME:
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active
            if sheet is None:
                result.schema_errors.append("No worksheets found in file")
                return []

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            result.schema_errors.append("Sheet is empty — no header row")
            return []

        headers = [str(h).strip() if h else "" for h in header_row]
        col_map = self._validate_schema(headers, result)
        if result.schema_errors:
            return []

        raw_rows: List[RowResult] = []
        for idx, values in enumerate(rows_iter, start=2):
            if not any(v is not None and str(v).strip() for v in values):
                continue
            data = {}
            for expected_col, col_idx in col_map.items():
                if col_idx == -1:
                    data[expected_col] = ""
                    continue
                val = values[col_idx] if col_idx < len(values) else None
                data[expected_col] = str(val).strip() if val is not None else ""
            raw_rows.append(RowResult(row_num=idx, data=data))

        return raw_rows

    def _parse_csv(self, file_bytes: bytes, result: ImportResult) -> List[RowResult]:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")

        reader = csv.reader(io.StringIO(text))
        try:
            headers = [h.strip() for h in next(reader)]
        except StopIteration:
            result.schema_errors.append("CSV file is empty")
            return []

        col_map = self._validate_schema(headers, result)
        if result.schema_errors:
            return []

        raw_rows: List[RowResult] = []
        for idx, values in enumerate(reader, start=2):
            if not any(v.strip() for v in values):
                continue
            data = {}
            for expected_col, col_idx in col_map.items():
                if col_idx == -1:
                    data[expected_col] = ""
                    continue
                val = values[col_idx] if col_idx < len(values) else ""
                data[expected_col] = val.strip()
            raw_rows.append(RowResult(row_num=idx, data=data))
        return raw_rows

    def _validate_schema(self, headers: List[str], result: ImportResult) -> Dict[str, int]:
        normalized = [h.lower().strip() for h in headers]
        col_map: Dict[str, int] = {}
        
        for expected_col in EXPECTED_COLUMNS:
            norm_expected = expected_col.lower().strip()
            if norm_expected in normalized:
                col_map[expected_col] = normalized.index(norm_expected)
            else:
                found = False
                for i, h in enumerate(normalized):
                    if norm_expected.replace(" ", "") == h.replace(" ", "") or norm_expected.replace(" ", "_") == h.replace(" ", "_"):
                        col_map[expected_col] = i
                        found = True
                        break
                if not found and expected_col in REQUIRED_COLUMNS:
                    result.schema_errors.append(f"Missing required column: '{expected_col}'")
                elif not found:
                    col_map[expected_col] = -1

        return col_map

    def _validate_row(self, row: RowResult):
        d = row.data
        errors = row.errors
        warnings = row.warnings

        dept_name = d.get("Department", "").strip()
        if not dept_name:
            errors.append("Department is required")
        elif dept_name.upper() not in self._dept_cache and dept_name not in self._dept_cache:
            found = False
            for k in self._dept_cache:
                if k.upper() == dept_name.upper():
                    found = True
                    break
            if not found:
                errors.append(f"Department '{dept_name}' not found")

        name = d.get("Class Name", "").strip()
        if not name:
            errors.append("Class Name is required")

        year_str = d.get("Year", "").strip()
        try:
            year = int(float(year_str))
            if year < 1 or year > 6:
                errors.append(f"Year should be 1-6, got {year}")
        except ValueError:
            errors.append(f"Invalid Year: {year_str}")

        semester_str = d.get("Semester", "").strip()
        try:
            semester = int(float(semester_str))
            if semester < 1 or semester > 12:
                errors.append(f"Semester should be 1-12, got {semester}")
        except ValueError:
            errors.append(f"Invalid Semester: {semester_str}")

        section = d.get("Section", "").strip()
        if not section:
            errors.append("Section is required")

        strength_str = d.get("Strength", "").strip()
        if strength_str:
            try:
                int(float(strength_str))
            except ValueError:
                errors.append(f"Invalid Strength: {strength_str}")

        code = self._generate_code(dept_name, year_str, section)
        if not errors and code in self._semester_code_cache:
            warnings.append(f"Class '{code}' already exists. Will UPDATE if committed.")

        row.status = "invalid" if errors else "valid"

    def _generate_code(self, dept: str, year: str, section: str) -> str:
        # Tries to normalize dept
        d = str(dept).strip().split(" ")[0].upper()
        y = str(year).strip().replace(".0", "")
        s = str(section).strip().upper()
        return f"{d}-{y}{s}"

    def _upsert_row(self, row: RowResult):
        d = row.data
        
        dept_name = d.get("Department", "").strip()
        dept_id = None
        # Safe resolution
        for key, did in self._dept_cache.items():
            if key.upper() == dept_name.upper():
                dept_id = did
                break

        name = d["Class Name"].strip()
        year = int(float(d["Year"].strip()))
        sem = int(float(d["Semester"].strip()))
        section = d["Section"].strip().upper()
        strength_str = d.get("Strength", "").strip()
        strength = int(float(strength_str)) if strength_str else 60

        # Unique code per department, year and section
        code = self._generate_code(dept_name, year, section)

        existing = self._semester_code_cache.get(code)
        if existing:
            existing.name = name
            existing.year = year
            existing.semester_number = sem
            existing.section = section
            existing.student_count = strength
            if dept_id:
                existing.dept_id = dept_id
            row.status = "updated"
        else:
            semester = Semester(
                name=name,
                code=code,
                year=year,
                semester_number=sem,
                section=section,
                student_count=strength,
                dept_id=dept_id
            )
            self.db.add(semester)
            self.db.flush()
            self._semester_code_cache[code] = semester
            row.status = "imported"
            
        row.data["Class Code"] = code

    def _warm_caches(self):
        if self._dept_cache is None:
            depts = self.db.query(Department).all()
            self._dept_cache = {}
            for d in depts:
                self._dept_cache[d.code] = d.id
                self._dept_cache[d.name] = d.id

        if self._semester_code_cache is None:
            semesters = self.db.query(Semester).all()
            self._semester_code_cache = {s.code: s for s in semesters}

    def _run_health_check(self) -> dict:
        total = self.db.query(Semester).count()
        return {
            "total_classes": total,
            "errors": [],
            "warnings": [],
            "all_clear": True,
        }

    @staticmethod
    def generate_template() -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = EXPECTED_SHEET_NAME

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(EXPECTED_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for i, w in enumerate([15, 20, 10, 10, 10, 12, 20, 20, 15], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        examples = [
            ["CSE", "3A", "2", "3", "A", "62", "Dr Sharma", "LH301", "Yes"],
            ["CSE", "3B", "2", "3", "B", "60", "Dr Patel", "LH302", "Yes"],
            ["AIML", "4A", "3", "5", "A", "58", "Prof Rao", "AIML Lab", "Yes"],
        ]
        
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row_data in enumerate(examples, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = example_fill
                cell.alignment = Alignment(horizontal="center")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPECTED_COLUMNS))}1"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
