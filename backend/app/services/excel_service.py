"""
Excel Generation Service for Official College Timetable Export.
"""
from io import BytesIO
from typing import List, Dict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session, joinedload
from app.db.models import Allocation, Semester, Subject
from app.core.config import get_settings

settings = get_settings()

COLLEGE_NAME = "K.RAMAKRISHNAN COLLEGE OF TECHNOLOGY(Autonomous)"

class TimetableExcelService:
    def __init__(self, db: Session):
        self.db = db
        
    def _get_subject_mnemonic(self, subject: Subject) -> str:
        name = subject.name
        words = name.split()
        if len(words) == 1:
            return name[:4].upper()
        mnemonic = ''.join(w[0].upper() for w in words if len(w) > 2)[:4]
        return mnemonic if mnemonic else name[:3].upper()

    def _get_component_suffix(self, alloc) -> str:
        comp = getattr(alloc, "academic_component", None) or (
            alloc.component_type.value if getattr(alloc, "component_type", None) else "theory"
        )
        if comp == "lab": return "(P)"
        if comp == "tutorial": return "(T)"
        if comp == "project": return "(PRJ)"
        if comp == "report": return "(RPT)"
        if comp == "self_study": return "(SS)"
        if comp == "seminar": return "(SEM)"
        return "(L)"

    def _get_semester_allocations(self, semester_id: int) -> tuple:
        allocations = self.db.query(Allocation).options(
            joinedload(Allocation.teacher),
            joinedload(Allocation.subject),
            joinedload(Allocation.room)
        ).filter(Allocation.semester_id == semester_id).all()
        
        grid = {}
        for day in range(5):
            grid[day] = {}
            for slot in range(settings.SLOTS_PER_DAY):
                grid[day][slot] = []
        
        for alloc in allocations:
            grid[alloc.day][alloc.slot].append(alloc)
            
        return grid, allocations

    def _build_semester_sheet(self, ws, semester: Semester, grid: Dict, allocations: List):
        widths = [8, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        dept_name = "DEPARTMENT OF ARTIFICIAL INTELLIGENCE & MACHINE LEARNING"
        if semester.department:
            dept_name = "DEPARTMENT OF " + semester.department.name.upper()
            
        current_year = datetime.now().year
        academic_year_str = f"{current_year}-{str(current_year+1)[-2:]}"
        semester_type_str = "EVEN SEMESTER" if semester.semester_number % 2 == 0 else "ODD SEMESTER"
        
        def to_roman(num):
            val = [10, 9, 5, 4, 1]
            syb = ["X", "IX", "V", "IV", "I"]
            roman_num = ''
            i = 0
            while num > 0:
                for _ in range(num // val[i]):
                    roman_num += syb[i]
                    num -= val[i]
                i += 1
            return roman_num
        
        roman_sem = to_roman(semester.semester_number)
        title = f"CLASS TIME TABLE - {roman_sem} SEMESTER - ACADEMIC YEAR {academic_year_str} ({semester_type_str})"

        ws.merge_cells('B1:I1')
        ws['A1'] = "***"
        ws['B1'] = COLLEGE_NAME
        ws['J1'] = "Format No:CPS-01"
        ws['B1'].font = Font(name='Times New Roman', size=14, bold=True)
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('B2:I2')
        ws['A2'] = "REVISION"
        ws['B2'] = dept_name
        ws['J2'] = "Issue No: 01"
        ws['B2'].font = Font(name='Times New Roman', size=11, bold=True)
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('B3:I3')
        ws['A3'] = "DATE"
        ws['B3'] = title
        ws['J3'] = f"Date: {datetime.now().strftime('%d.%m.%y')}"
        ws['B3'].font = Font(name='Times New Roman', size=12, bold=True)
        ws['B3'].alignment = Alignment(horizontal='center', vertical='center')

        room_counts = {}
        for alloc in allocations:
            if alloc.room and alloc.component_type.value == "theory":
                room_counts[alloc.room.name] = room_counts.get(alloc.room.name, 0) + 1
        primary_room = max(room_counts.items(), key=lambda x: x[1])[0] if room_counts else "***"
        section = semester.section or "A"
        strength = str(semester.student_count) if semester.student_count else "60"

        fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('B5:C5')
        ws['A5'] = "HOD"
        ws['B5'] = "***"
        ws['D5'] = "SECTION"
        ws['E5'] = section
        ws.merge_cells('F5:G5')
        ws['F5'] = "CHAIR PERSON"
        ws.merge_cells('H5:I5')
        ws['H5'] = "***"
        ws['J5'] = "ROOM NO."
        ws['K5'] = primary_room
        
        ws.merge_cells('B6:C6')
        ws['A6'] = "CLASS ADVISOR"
        ws['B6'] = "***"
        ws['D6'] = "STRENGTH"
        ws['E6'] = strength
        ws.merge_cells('F6:G6')
        ws['F6'] = "ASST.CLASS ADVISOR"
        ws.merge_cells('H6:I6')
        ws['H6'] = "***"
        ws['J6'] = "CLASS REP"
        ws['K6'] = "***"

        for row in [5, 6]:
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_blue
                cell.border = border_thin
                cell.font = Font(name='Times New Roman', size=9, bold=(col in [1, 4, 6, 10]))
                cell.alignment = Alignment(horizontal='left', vertical='center')

        headers = ["DAYS", "1", "2", "BREAK", "3", "LUNCH", "4", "5", "BREAK", "6", "7"]
        timings = ["8:45 a.m. -\n9:45 a.m.", "9:45 a.m. -\n10:45 a.m.", "10:45 a.m.-\n11:00 a.m.", "11:00 a.m.-\n12:00 p.m.", "12:00 p.m.-\n01:00 p.m.", "01:00 p.m.-\n02:00 p.m.", "02:00 p.m.-\n02:50 p.m.", "02:50 p.m.-\n03:05 p.m.", "03:05 p.m.-\n03:55 p.m.", "03:55 p.m.-\n04:45 p.m."]

        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=i)
            cell.value = header
            cell.font = Font(name='Times New Roman', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws.cell(row=9, column=1, value="").border = border_thin
        ws.cell(row=9, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        col_idx = 2
        for t in timings:
            cell = ws.cell(row=9, column=col_idx)
            cell.value = t
            cell.font = Font(name='Times New Roman', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_thin
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            col_idx += 1

        day_names = ["MON", "TUE", "WED", "THU", "FRI"]
        start_row = 10
        
        fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        fill_red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        fill_orange = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        fill_grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        break1_col = 4
        lunch_col = 6
        break2_col = 9

        for i in range(5):
            ws.cell(row=start_row+i, column=break1_col).fill = fill_grey
            ws.cell(row=start_row+i, column=break1_col).border = border_thin
            ws.cell(row=start_row+i, column=lunch_col).fill = fill_grey
            ws.cell(row=start_row+i, column=lunch_col).border = border_thin
            ws.cell(row=start_row+i, column=break2_col).fill = fill_grey
            ws.cell(row=start_row+i, column=break2_col).border = border_thin
            
        ws.merge_cells(start_row=start_row, start_column=break1_col, end_row=start_row+4, end_column=break1_col)
        ws.cell(row=start_row, column=break1_col, value="B\nR\nE\nA\nK").alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
        ws.cell(row=start_row, column=break1_col).font = Font(name='Times New Roman', bold=True)
        
        ws.merge_cells(start_row=start_row, start_column=lunch_col, end_row=start_row+4, end_column=lunch_col)
        ws.cell(row=start_row, column=lunch_col, value="L\nU\nN\nC\nH").alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
        ws.cell(row=start_row, column=lunch_col).font = Font(name='Times New Roman', bold=True)

        ws.merge_cells(start_row=start_row, start_column=break2_col, end_row=start_row+4, end_column=break2_col)
        ws.cell(row=start_row, column=break2_col, value="B\nR\nE\nA\nK").alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
        ws.cell(row=start_row, column=break2_col).font = Font(name='Times New Roman', bold=True)

        for day in range(5):
            row_num = start_row + day
            cell = ws.cell(row=row_num, column=1, value=day_names[day])
            cell.font = Font(name='Times New Roman', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
            cell.fill = fill_grey

            slot_to_col = {0: 2, 1: 3, 2: 5, 3: 7, 4: 8, 5: 10, 6: 11}
            skip_cols = set()
            
            for slot in range(7):
                col_num = slot_to_col[slot]
                if col_num in skip_cols:
                    continue
                    
                allocs = grid[day].get(slot, [])
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border_thin
                cell.font = Font(name='Times New Roman', size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                if not allocs:
                    cell.value = "-"
                    continue
                    
                is_lab = any(a.component_type.value == "lab" for a in allocs)
                is_elective = any(a.is_elective for a in allocs)
                
                if is_lab:
                    cell.fill = fill_red
                    texts = []
                    for a in allocs:
                        texts.append(f"{self._get_subject_mnemonic(a.subject)}{self._get_component_suffix(a)}-{a.teacher.name[:10] if a.teacher else '***'}")
                    cell.value = " / ".join(texts)
                    
                    if slot < 6:
                        next_allocs = grid[day].get(slot+1, [])
                        next_is_same_lab = any(a.component_type.value == "lab" for a in next_allocs)
                        if next_is_same_lab:
                            next_col = slot_to_col[slot+1]
                            ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=next_col)
                            ws.cell(row=row_num, column=next_col).border = border_thin
                            skip_cols.add(next_col)
                else:
                    if is_elective:
                        cell.fill = fill_orange
                    else:
                        cell.fill = fill_yellow
                        
                    if len(allocs) > 1:
                        texts = []
                        for a in allocs:
                            texts.append(f"{self._get_subject_mnemonic(a.subject)}{self._get_component_suffix(a)}-{a.teacher.name[:10] if a.teacher else '***'}")
                        cell.value = " / ".join(texts)
                    else:
                        alloc = allocs[0]
                        subj_text = f"{self._get_subject_mnemonic(alloc.subject)}{self._get_component_suffix(alloc)}"
                        tname = alloc.teacher.name if alloc.teacher else "***"
                        if len(tname) > 15:
                            parts = tname.split()
                            if len(parts) > 1:
                                tname = f"{parts[0]} {parts[1][:1]}."
                        cell.value = f"{subj_text}\n{tname}"

        start_row_sub = start_row + 6
        
        col_mappings = [
            (1, 2, "SUB CODE"),
            (3, 5, "SUBJECT NAME"),
            (6, 6, "MNEMONIC"),
            (7, 7, "CREDIT"),
            (8, 9, "STAFF NAME(M)"),
            (10, 10, "DEPT"),
            (11, 11, "TOTAL HOURS")
        ]
        
        for start_c, end_c, text in col_mappings:
            if start_c != end_c:
                ws.merge_cells(start_row=start_row_sub, start_column=start_c, end_row=start_row_sub, end_column=end_c)
            cell = ws.cell(row=start_row_sub, column=start_c, value=text)
            cell.font = Font(name='Times New Roman', size=10, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_blue
            for c in range(start_c, end_c + 1):
                ws.cell(row=start_row_sub, column=c).border = border_thin
                
        subject_data = {}
        for alloc in allocations:
            key = alloc.subject_id
            if key not in subject_data:
                subject_data[key] = {
                    "subject": alloc.subject,
                    "teachers": set(),
                    "component": alloc.component_type,
                    "slots": set()
                }
            if alloc.teacher:
                subject_data[key]["teachers"].add(alloc.teacher.name)
            subject_data[key]["slots"].add((alloc.day, alloc.slot))

        curr_row = start_row_sub + 1
        for data in subject_data.values():
            subj = data["subject"]
            tname = ", ".join(data["teachers"]) if data["teachers"] else "***"
            hours = len(data["slots"])
            comp_type = data["component"].value.upper()
            if comp_type == "LAB": comp_type = "PRACTICAL"
            elif comp_type == "THEORY": comp_type = "THEORY"
            
            # Create a dummy allocation for mnemonic helper
            dummy_alloc = type('Dummy', (), {'component_type': data['component'], 'academic_component': None})()
            mnemonic = f"{self._get_subject_mnemonic(subj)}{self._get_component_suffix(dummy_alloc)}"
            dept = semester.department.code if semester.department else "AI"
            
            row_data = [
                (1, 2, subj.code),
                (3, 5, f"{subj.name} ({comp_type})"),
                (6, 6, mnemonic),
                (7, 7, "3"),
                (8, 9, tname),
                (10, 10, dept),
                (11, 11, str(hours))
            ]
            
            for start_c, end_c, val in row_data:
                if start_c != end_c:
                    ws.merge_cells(start_row=curr_row, start_column=start_c, end_row=curr_row, end_column=end_c)
                cell = ws.cell(row=curr_row, column=start_c, value=val)
                cell.font = Font(name='Times New Roman', size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                for c in range(start_c, end_c + 1):
                    ws.cell(row=curr_row, column=c).border = border_thin
                    
            curr_row += 1

        curr_row += 3
        ws.cell(row=curr_row, column=2, value="Dept.T.T. Coordinator").font = Font(name='Times New Roman', bold=True)
        ws.cell(row=curr_row, column=5, value="HoD-AI").font = Font(name='Times New Roman', bold=True)
        ws.cell(row=curr_row, column=8, value="CoT").font = Font(name='Times New Roman', bold=True)
        ws.cell(row=curr_row, column=10, value="HAA").font = Font(name='Times New Roman', bold=True)
        
    def generate_semester_excel(self, semester_id: int) -> BytesIO:
        semester = self.db.query(Semester).options(
            joinedload(Semester.department)
        ).filter(Semester.id == semester_id).first()
        
        if not semester:
            raise ValueError(f"Semester {semester_id} not found")
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = semester.code[:30]
        
        grid, allocations = self._get_semester_allocations(semester_id)
        self._build_semester_sheet(ws, semester, grid, allocations)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_all_timetables_excel(self) -> BytesIO:
        semesters = self.db.query(Semester).options(
            joinedload(Semester.department)
        ).order_by(Semester.year, Semester.code).all()
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        for sem in semesters:
            grid, allocations = self._get_semester_allocations(sem.id)
            if not allocations:
                continue
                
            safe_title = sem.code[:31]
            suffix = 1
            original_title = safe_title
            while safe_title in wb.sheetnames:
                safe_title = f"{original_title[:28]}_{suffix}"
                suffix += 1
                
            ws = wb.create_sheet(title=safe_title)
            self._build_semester_sheet(ws, sem, grid, allocations)
            
        if not wb.sheetnames:
            wb.create_sheet("Empty")
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
