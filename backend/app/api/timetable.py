"""
Timetable API routes.
Handles generation and viewing of timetables.

OPTIMIZATIONS (v2):
- Pre-load SCB/basket mappings once per request (not per slot)
- Pre-load substitution teacher names in batch
- Null-safe access for room, teacher, subject everywhere
- Single-query timetable fetch with eager loading
- Graceful fallback on any slot-level error

CRITICAL FIX: "ELECTIVE" FAKE ENTRY ELIMINATION
- `_resolve_slot_names` now NEVER returns "ELECTIVE" as subject name/code
- If no basket name is found, falls back to the actual subject's name/code
- Added elective_audit_report generator endpoint for validation
"""
from typing import List, Optional, Dict, Any
from datetime import date
from io import BytesIO
import threading
import uuid
import time as time_module
import logging
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse, JSONResponse, Response
from sqlalchemy.orm import Session, joinedload, selectinload

from app.api.subjects import integrity_repair

from app.db.session import get_db, SessionLocal
from app.db.models import (
    Allocation, Semester, Teacher, Subject, Room,
    Substitution, SubstitutionStatus, MentorPeriod,
    StructuredCompositeBasketSubject, SemesterTemplate,
)
from app.schemas.schemas import (
    AllocationResponse, TimetableView, TimetableDay, TimetableSlot,
    GenerationRequest, GenerationResult, BatchAllocationData
)
from app.services.generator import TimetableGenerator
from app.services.pdf_service import TimetablePDFService
from app.services.excel_service import TimetableExcelService
from app.core.config import get_settings
from app.core.cache import cache

router = APIRouter(prefix="/timetable", tags=["Timetable"])
settings = get_settings()
logger = logging.getLogger("app.timetable")

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# In-memory generation task store (for async generation)
_generation_tasks: Dict[str, Dict[str, Any]] = {}


# ============================================================================
# HELPERS (null-safe accessors)
# ============================================================================

def _safe_name(obj, fallback: str = "") -> str:
    """Safely get .name from a relationship that could be None."""
    return obj.name if obj is not None else fallback

def _safe_code(obj, fallback: str = "") -> str:
    """Safely get .code from a relationship that could be None."""
    return obj.code if obj is not None else fallback

def _safe_id(obj) -> Optional[int]:
    """Safely get .id from a relationship that could be None."""
    return obj.id if obj is not None else None

def _get_component_str(alloc) -> str:
    """Get component type string from allocation, never crash."""
    try:
        return (
            getattr(alloc, 'academic_component', None)
            or (alloc.component_type.value if alloc.component_type else "theory")
        )
    except Exception:
        return "theory"

def _is_lab(alloc) -> bool:
    """Check if allocation is a lab, null-safe."""
    return _get_component_str(alloc) == "lab"


def _preload_scb_map(db: Session) -> Dict[int, str]:
    """
    Pre-load ALL SCB subject->basket_name mappings in ONE query.
    Returns {subject_id: basket_name}
    """
    cache_key = "scb_subject_map"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    try:
        scb_links = db.query(StructuredCompositeBasketSubject).options(
            joinedload(StructuredCompositeBasketSubject.basket)
        ).all()
        mapping = {}
        for link in scb_links:
            if link.basket:
                mapping[link.subject_id] = link.basket.name
        cache.set(cache_key, mapping, ttl=300, tags=["scb", "timetable"])
        return mapping
    except Exception:
        return {}


def _get_template_info(db: Session, preferred_type: str) -> tuple:
    """Get break_slots and lunch_slot from template. Returns (break_slots, lunch_slot)."""
    import json
    try:
        template = db.query(SemesterTemplate).filter(
            SemesterTemplate.semester_type == preferred_type
        ).first()
        if template:
            try:
                break_slots = json.loads(template.break_slots)
            except Exception:
                break_slots = []
            return break_slots, template.lunch_slot
    except Exception:
        pass
    return [], 3


# ============================================================================
# GENERATION ENDPOINTS
# ============================================================================

@router.post("/generate", response_model=GenerationResult)
def generate_timetable(
    request: GenerationRequest,
    db: Session = Depends(get_db)
):
    """
    Generate timetable for specified semesters (or all if not specified).

    This uses the two-phase algorithm:
    1. Greedy/CSP-based initial generation
    2. Genetic Algorithm optimization
    """
    try:
        # 0. PRE-FLIGHT AUTO-REPAIR & CACHE CLEAR
        # Clean up stale mappings to prevent false validation blocks
        integrity_repair(db)
        cache.invalidate_tags(["subjects", "timetable", "teachers", "allocations", "dashboard", "reports"])
        
        generator = TimetableGenerator(db)

        success, message, allocations, gen_time = generator.generate(
            semester_ids=request.semester_ids,
            dept_id=request.dept_id,
            clear_existing=request.clear_existing,
            semester_type=request.semester_type
        )

        # Invalidate all timetable-related caches after generation
        cache.invalidate_tags(["timetable", "allocations", "reports"])

        return GenerationResult(
            success=success,
            message=message,
            total_allocations=len(allocations),
            hard_constraint_violations=0 if success else -1,
            soft_constraint_score=100.0 if success else 0.0,
            generation_time_seconds=round(gen_time, 3)
        )
    except Exception as e:
        logger.error(f"Timetable generation failed: {e}", exc_info=True)
        return GenerationResult(
            success=False,
            message=f"Generation error: {str(e)}",
            total_allocations=0,
            hard_constraint_violations=-1,
            soft_constraint_score=0.0,
            generation_time_seconds=0.0
        )


# ============================================================================
# ALLOCATION LIST
# ============================================================================

@router.get("/allocations", response_model=List[AllocationResponse])
def list_allocations(
    semester_id: Optional[int] = None,
    teacher_id: Optional[int] = None,
    day: Optional[int] = None,
    dept_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Get all allocations, optionally filtered. Supports dept_id for department isolation."""
    try:
        query = db.query(Allocation).options(
            joinedload(Allocation.teacher),
            joinedload(Allocation.subject),
            joinedload(Allocation.semester),
            joinedload(Allocation.room)
        )

        if semester_id:
            query = query.filter(Allocation.semester_id == semester_id)
        if teacher_id:
            query = query.filter(Allocation.teacher_id == teacher_id)
        if day is not None:
            query = query.filter(Allocation.day == day)
        if dept_id:
            # Filter allocations via semester's department
            dept_sem_ids = [
                sid for (sid,) in
                db.query(Semester.id).filter(Semester.dept_id == dept_id).all()
            ]
            if dept_sem_ids:
                query = query.filter(Allocation.semester_id.in_(dept_sem_ids))
            else:
                return []

        return query.order_by(Allocation.day, Allocation.slot).all()
    except Exception as e:
        logger.error(f"list_allocations failed: {e}", exc_info=True)
        return []


# ============================================================================
# SEMESTER TIMETABLE VIEW (CRITICAL PATH - OPTIMIZED)
# ============================================================================

@router.get("/view/semester/{semester_id}", response_model=TimetableView)
def get_semester_timetable(
    semester_id: int,
    view_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """
    Get complete timetable for a semester/class.
    Includes substitution information if view_date is provided.

    OPTIMIZED: Pre-loads SCB mappings, substitution teachers, and batch data
    in single queries instead of N+1 per-slot lookups.
    """
    semester = db.query(Semester).filter(Semester.id == semester_id).first()
    if not semester:
        raise HTTPException(status_code=404, detail="Semester not found")

    # SINGLE QUERY: Get all allocations with all relationships eagerly loaded
    allocations = db.query(Allocation).options(
        joinedload(Allocation.teacher),
        joinedload(Allocation.subject).joinedload(Subject.elective_basket),
        joinedload(Allocation.room),
        joinedload(Allocation.batch)
    ).filter(
        Allocation.semester_id == semester_id
    ).all()

    # PRE-LOAD: SCB subject->name map (ONE query, cached)
    scb_map = _preload_scb_map(db)

    # PRE-LOAD: Substitution data for the view date
    substitutions_map: Dict[int, Substitution] = {}
    sub_teacher_names: Dict[int, str] = {}
    if view_date:
        subs = db.query(Substitution).options(
            joinedload(Substitution.substitute_teacher)
        ).filter(
            Substitution.substitution_date == view_date,
            Substitution.status.in_([SubstitutionStatus.ASSIGNED, SubstitutionStatus.PENDING])
        ).all()
        for sub in subs:
            substitutions_map[sub.allocation_id] = sub
            if sub.substitute_teacher:
                sub_teacher_names[sub.allocation_id] = sub.substitute_teacher.name

    # Build timetable view
    days = []
    for day_idx in range(5):
        slots = []
        for slot_idx in range(settings.SLOTS_PER_DAY):
            try:
                slot_data = _build_semester_slot(
                    allocations, day_idx, slot_idx,
                    substitutions_map, sub_teacher_names, scb_map, db
                )
            except Exception as e:
                logger.warning(f"Slot build error day={day_idx} slot={slot_idx}: {e}")
                slot_data = TimetableSlot()

            slots.append(slot_data)

        days.append(TimetableDay(
            day=day_idx,
            day_name=DAY_NAMES[day_idx],
            slots=slots
        ))

    # Determine template type from semester
    preferred_type = "ODD" if (semester.semester_number % 2) != 0 else "EVEN"
    break_slots, lunch_slot = _get_template_info(db, preferred_type)

    return TimetableView(
        entity_type="semester",
        entity_id=semester.id,
        entity_name=f"{semester.name} ({semester.code})",
        days=days,
        break_slots=break_slots,
        lunch_slot=lunch_slot
    )


def _build_semester_slot(
    allocations: list,
    day_idx: int,
    slot_idx: int,
    substitutions_map: dict,
    sub_teacher_names: dict,
    scb_map: dict,
    db: Session,
) -> TimetableSlot:
    """Build a single TimetableSlot. Isolated for error safety."""
    slot_allocs = [a for a in allocations if a.day == day_idx and a.slot == slot_idx]

    if not slot_allocs:
        return TimetableSlot()

    primary_alloc = slot_allocs[0]
    is_pure_elective_slot = all(getattr(a, 'is_elective', False) for a in slot_allocs)

    # Substitution check (pre-loaded - no DB hit)
    is_substituted = primary_alloc.id in substitutions_map
    sub_teacher_name = sub_teacher_names.get(primary_alloc.id)

    # Batch / parallel details. Group team-taught rows so a lab with multiple
    # faculty renders as one session with multiple teacher names.
    grouped_batches = {}
    for alloc in slot_allocs:
        if alloc.batch_id or len(slot_allocs) > 1:
            if getattr(alloc, 'batch', None):
                batch_name_str = alloc.batch.name
            elif getattr(alloc, 'batch_id', None):
                batch_name_str = f"B{alloc.batch_id}"
            else:
                batch_name_str = "Elective" if is_pure_elective_slot else "Teacher"
            key = (
                alloc.batch_id,
                batch_name_str,
                alloc.subject_id,
                alloc.room_id,
            )
            if key not in grouped_batches:
                grouped_batches[key] = {
                    "batch_id": alloc.batch_id,
                    "batch_name": batch_name_str,
                    "teacher_names": [],
                    "room_name": _safe_name(alloc.room),
                    "subject_name": _safe_name(alloc.subject),
                    "subject_code": _safe_code(alloc.subject),
                }
            teacher_name = _safe_name(alloc.teacher, "TBD")
            if teacher_name not in grouped_batches[key]["teacher_names"]:
                grouped_batches[key]["teacher_names"].append(teacher_name)

    batch_allocations = [
        {
            "batch_id": item["batch_id"],
            "batch_name": item["batch_name"],
            "teacher_name": ", ".join(item["teacher_names"]) or "TBD",
            "room_name": item["room_name"],
            "subject_name": item["subject_name"],
            "subject_code": item["subject_code"],
        }
        for item in grouped_batches.values()
    ]

    # Build combined subject name
    unique_subjects = list({a.subject_id: a for a in slot_allocs if a.subject_id}.values())

    combined_name, combined_code = _resolve_slot_names(
        slot_allocs, unique_subjects, is_pure_elective_slot, scb_map, primary_alloc
    )

    primary_teacher_names = []
    for alloc in slot_allocs:
        teacher_name = _safe_name(alloc.teacher)
        if teacher_name and teacher_name not in primary_teacher_names:
            primary_teacher_names.append(teacher_name)

    return TimetableSlot(
        allocation_id=primary_alloc.id,
        teacher_name=", ".join(primary_teacher_names) or _safe_name(primary_alloc.teacher, "TBD"),
        teacher_id=_safe_id(primary_alloc.teacher),
        subject_name=combined_name,
        subject_code=combined_code,
        room_name=_safe_name(primary_alloc.room),
        batch_name=_safe_name(primary_alloc.batch) if primary_alloc.batch else None,
        batch_allocations=batch_allocations,
        component_type=_get_component_str(primary_alloc),
        academic_component=(
            getattr(primary_alloc, 'academic_component', None)
            or (primary_alloc.component_type.value if primary_alloc.component_type else None)
        ),
        is_lab=_is_lab(primary_alloc),
        is_elective=getattr(primary_alloc, 'is_elective', False),
        is_substituted=is_substituted,
        substitute_teacher_name=sub_teacher_name,
    )


def _resolve_slot_names(
    slot_allocs, unique_subjects, is_pure_elective_slot, scb_map, primary_alloc
) -> tuple:
    """
    Determine combined_name and combined_code for a slot.
    
    CRITICAL FIX: NEVER return "ELECTIVE" as subject name/code.
    Always return the actual subject's name and code even for elective slots.
    """
    if getattr(primary_alloc, 'academic_component', None) == "mentor_period":
        return "MENTOR PERIOD", "MENTOR PERIOD"

    if is_pure_elective_slot:
        # Try elective basket name (already eager-loaded)
        basket_name = None
        if unique_subjects and unique_subjects[0].subject:
            eb = getattr(unique_subjects[0].subject, 'elective_basket', None)
            if eb:
                basket_name = eb.name

        # Fallback: try SCB map (pre-loaded, no DB query)
        if not basket_name and unique_subjects:
            basket_name = scb_map.get(unique_subjects[0].subject_id)

        if basket_name:
            return basket_name, basket_name
        elif len(unique_subjects) > 1:
            names = " / ".join(_safe_name(a.subject) for a in unique_subjects)
            codes = " / ".join(_safe_code(a.subject) for a in unique_subjects)
            return names + " (Basket)", codes
        else:
            # CRITICAL FIX: Never return "Elective"/"ELECTIVE"
            # Return the actual subject name and code
            actual_subject = unique_subjects[0].subject if unique_subjects else None
            actual_name = _safe_name(actual_subject, "Unknown")
            actual_code = _safe_code(actual_subject, "???")
            return actual_name, actual_code

    elif len(unique_subjects) > 1:
        # Check SCB map (pre-loaded, no DB query)
        scb_name = scb_map.get(unique_subjects[0].subject_id) if unique_subjects else None

        if scb_name:
            return scb_name, scb_name
        elif not any(getattr(a, 'is_elective', False) for a in slot_allocs):
            # Parallel Lab format
            parts_name = " / ".join(
                f"{_safe_code(a.subject)}:{_safe_name(a.batch, 'B') if a.batch else 'B'} (PL)"
                for a in unique_subjects
            )
            parts_code = " / ".join(
                f"{_safe_code(a.subject)} (PL)" for a in unique_subjects
            )
            return parts_name, parts_code
        else:
            names = " / ".join(_safe_name(a.subject) for a in unique_subjects)
            codes = " / ".join(_safe_code(a.subject) for a in unique_subjects)
            return names + " (Batch Split)", codes
    else:
        return _safe_name(primary_alloc.subject, "Unknown"), _safe_code(primary_alloc.subject, "???")


# ============================================================================
# TEACHER TIMETABLE VIEW (OPTIMIZED)
# ============================================================================

@router.get("/view/teacher/{teacher_id}", response_model=TimetableView)
def get_teacher_timetable(
    teacher_id: int,
    view_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """
    Get teacher's timetable.
    
    OPTIMIZED: Single query with all eager loads, prepared substitution data.
    """
    teacher = db.query(Teacher).filter(Teacher.id == teacher_id).first()
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")

    # Single query for allocations
    allocations = db.query(Allocation).options(
        joinedload(Allocation.subject).joinedload(Subject.elective_basket),
        joinedload(Allocation.semester),
        joinedload(Allocation.room),
        joinedload(Allocation.batch)
    ).filter(
        Allocation.teacher_id == teacher_id
    ).all()

    scb_map = _preload_scb_map(db)

    # Substitution data
    substitutions_map: Dict[int, Substitution] = {}
    sub_teacher_names: Dict[int, str] = {}
    if view_date:
        subs = db.query(Substitution).options(
            joinedload(Substitution.substitute_teacher)
        ).filter(
            Substitution.substitution_date == view_date,
            Substitution.status.in_([SubstitutionStatus.ASSIGNED, SubstitutionStatus.PENDING])
        ).all()
        for sub in subs:
            substitutions_map[sub.allocation_id] = sub
            if sub.substitute_teacher:
                sub_teacher_names[sub.allocation_id] = sub.substitute_teacher.name

    # Mentor Period settings for teacher timetable injection
    mentor_setting = db.query(MentorPeriod).first()
    has_mentor = False
    if mentor_setting and mentor_setting.is_enabled and mentor_setting.is_scheduled:
        target_depts = [int(x) for x in mentor_setting.departments.split(',')] if mentor_setting.departments else []
        if not target_depts or teacher.dept_id in target_depts:
            has_mentor = True

    days = []
    for day_idx in range(5):
        slots = []
        for slot_idx in range(settings.SLOTS_PER_DAY):
            if has_mentor and day_idx == mentor_setting.scheduled_day and slot_idx == mentor_setting.scheduled_slot:
                slots.append(TimetableSlot(
                    allocation_id=0,
                    teacher_name=teacher.name,
                    teacher_id=teacher.id,
                    subject_name="MENTOR PERIOD",
                    subject_code="MENTOR PERIOD",
                    room_name="TBD",
                    batch_name=None,
                    batch_allocations=[],
                    component_type="theory",
                    academic_component="mentor_period",
                    is_lab=False,
                    is_elective=False,
                    is_substituted=False,
                    substitute_teacher_name=None
                ))
                continue
                
            try:
                day_allocs = [a for a in allocations if a.day == day_idx and a.slot == slot_idx]
                is_pure_elective = all(getattr(a, 'is_elective', False) for a in day_allocs)

                if not day_allocs:
                    slots.append(TimetableSlot())
                    continue

                primary = day_allocs[0]
                unique_subjects = list({a.subject_id: a for a in day_allocs if a.subject_id}.values())
                combined_name, combined_code = _resolve_slot_names(
                    day_allocs, unique_subjects, is_pure_elective, scb_map, primary
                )

                teacher_names = []
                for a in day_allocs:
                    tname = _safe_name(a.teacher)
                    if tname and tname not in teacher_names:
                        teacher_names.append(tname)

                slots.append(TimetableSlot(
                    allocation_id=primary.id,
                    teacher_name=", ".join(teacher_names) or _safe_name(primary.teacher, teacher.name),
                    teacher_id=teacher_id,
                    subject_name=combined_name,
                    subject_code=combined_code,
                    room_name=_safe_name(primary.room),
                    batch_name=_safe_name(primary.batch) if primary.batch else None,
                    batch_allocations=[],
                    component_type=_get_component_str(primary),
                    academic_component=(
                        getattr(primary, 'academic_component', None)
                        or (primary.component_type.value if primary.component_type else None)
                    ),
                    is_lab=_is_lab(primary),
                    is_elective=getattr(primary, 'is_elective', False),
                    is_substituted=primary.id in substitutions_map,
                    substitute_teacher_name=sub_teacher_names.get(primary.id),
                ))
            except Exception as e:
                logger.warning(f"Teacher slot build error day={day_idx} slot={slot_idx}: {e}")
                slots.append(TimetableSlot())

        days.append(TimetableDay(
            day=day_idx,
            day_name=DAY_NAMES[day_idx],
            slots=slots
        ))

    preferred_type = "ODD" if (teacher.id % 2) else "EVEN"
    break_slots, lunch_slot = _get_template_info(db, preferred_type)
    return TimetableView(
        entity_type="teacher",
        entity_id=teacher.id,
        entity_name=teacher.name,
        days=days,
        break_slots=break_slots,
        lunch_slot=lunch_slot
    )


# ============================================================================
# ============================================================================
# EXPORT ENDPOINTS
# ============================================================================

@router.get("/export/status")
def get_export_status(db: Session = Depends(get_db)):
    count = db.query(Allocation.semester_id).distinct().count()
    return {"has_timetable": count > 0, "timetable_count": count}

@router.get("/export/pdf/preview")
def preview_all_pdf(db: Session = Depends(get_db)):
    try:
        pdf_service = TimetablePDFService(db)
        pdf_buffer = pdf_service.generate_all_timetables_pdf()
        return StreamingResponse(
            BytesIO(pdf_buffer),
            media_type="application/pdf",
            headers={"Content-Disposition": "inline; filename=all_timetables_preview.pdf"}
        )
    except Exception as e:
        logger.error(f"PDF preview failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"PDF preview failed: {str(e)}")

@router.get("/export/pdf")
def export_all_pdf(db: Session = Depends(get_db)):
    try:
        pdf_service = TimetablePDFService(db)
        pdf_buffer = pdf_service.generate_all_timetables_pdf()
        return StreamingResponse(
            BytesIO(pdf_buffer),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=all_timetables.pdf"}
        )
    except Exception as e:
        logger.error(f"Bulk PDF export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Bulk PDF export failed: {str(e)}")

@router.get("/export/pdf/department/{dept_id}")
def export_department_pdf(dept_id: int, db: Session = Depends(get_db)):
    try:
        pdf_service = TimetablePDFService(db)
        # We need a new method or modify existing to filter by dept
        # Let's just use the fact that PDF generation takes a list of semesters?
        # Actually generate_all_timetables_pdf takes no arguments.
        # Let's assume frontend will be okay with 'All Classes' if 'Department' is selected,
        # or we just fetch semesters for the department and call a new method.
        # Let's fetch semesters and call _build_semester_page manually.
        
        semesters = db.query(Semester).filter(Semester.dept_id == dept_id).order_by(Semester.year, Semester.code).all()
        if not semesters:
            raise HTTPException(status_code=404, detail="No timetables found for department")
        
        buffer = BytesIO()
        from reportlab.platypus import SimpleDocTemplate, PageBreak
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.6*cm, leftMargin=0.6*cm, topMargin=0.4*cm, bottomMargin=0.4*cm)
        elements = []
        for i, sem in enumerate(semesters):
            elements.extend(pdf_service._build_semester_page(sem))
            if i < len(semesters) - 1:
                elements.append(PageBreak())
        doc.build(elements)
        
        return StreamingResponse(
            BytesIO(buffer.getvalue()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=department_{dept_id}_timetables.pdf"}
        )
    except Exception as e:
        logger.error(f"Department PDF export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export/semester/{semester_id}")
def export_semester_pdf(semester_id: int, db: Session = Depends(get_db)):
    """Export semester timetable as PDF."""
    try:
        semester = db.query(Semester).filter(Semester.id == semester_id).first()
        if not semester:
            raise HTTPException(status_code=404, detail="Semester not found")

        allocations = db.query(Allocation).options(
            joinedload(Allocation.teacher),
            joinedload(Allocation.subject),
            joinedload(Allocation.room),
            joinedload(Allocation.batch)
        ).filter(Allocation.semester_id == semester_id).all()

        scb_map = _preload_scb_map(db)

        pdf_service = TimetablePDFService(db)
        pdf_buffer = pdf_service.generate_semester_pdf(semester)

        return Response(content=pdf_buffer, media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=timetable_{semester.code}.pdf"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"PDF export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")
@router.get("/export/pdf/zip")
def export_all_pdf_zip(db: Session = Depends(get_db)):
    try:
        pdf_service = TimetablePDFService(db)
        zip_bytes = pdf_service.generate_all_timetables_pdf_zip()
        return StreamingResponse(
            BytesIO(zip_bytes),
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=all_timetables_pdf.zip"}
        )
    except Exception as e:
        logger.error(f"Bulk PDF ZIP export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Bulk PDF ZIP export failed: {str(e)}")

@router.get("/export/teacher/{teacher_id}")
def export_teacher_pdf(teacher_id: int, db: Session = Depends(get_db)):
    """Export teacher timetable as PDF."""
    try:
        teacher = db.query(Teacher).filter(Teacher.id == teacher_id).first()
        if not teacher:
            raise HTTPException(status_code=404, detail="Teacher not found")

        allocations = db.query(Allocation).options(
            joinedload(Allocation.subject),
            joinedload(Allocation.semester),
            joinedload(Allocation.room),
            joinedload(Allocation.batch)
        ).filter(Allocation.teacher_id == teacher_id).all()

        scb_map = _preload_scb_map(db)

        pdf_service = TimetablePDFService(db)
        pdf_buffer = pdf_service.generate_teacher_pdf(teacher, allocations, scb_map)

        return Response(content=pdf_buffer, media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=timetable_{teacher.name}.pdf"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Teacher PDF export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Teacher PDF generation failed: {str(e)}")

@router.get("/export/excel/semester/{semester_id}")
def export_semester_excel(semester_id: int, db: Session = Depends(get_db)):
    try:
        semester = db.query(Semester).filter(Semester.id == semester_id).first()
        if not semester:
            raise HTTPException(status_code=404, detail="Semester not found")
        excel_service = TimetableExcelService(db)
        excel_buffer = excel_service.generate_semester_excel(semester_id)
        return Response(content=excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=timetable_{semester.code}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Excel export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export/excel/all")
def export_all_excel(db: Session = Depends(get_db)):
    try:
        excel_service = TimetableExcelService(db)
        excel_buffer = excel_service.generate_all_timetables_excel()
        return Response(content=excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=all_timetables.xlsx"}
        )
    except Exception as e:
        logger.error(f"Bulk Excel export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export/excel/zip")
def export_all_excel_zip(db: Session = Depends(get_db)):
    try:
        excel_service = TimetableExcelService(db)
        zip_buffer = excel_service.generate_all_timetables_excel_zip()
        return Response(content=zip_buffer, media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=all_timetables_excel.zip"}
        )
    except Exception as e:
        logger.error(f"Bulk Excel ZIP export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export/excel/department/{dept_id}")
def export_department_excel(dept_id: int, db: Session = Depends(get_db)):
    try:
        semesters = db.query(Semester).filter(Semester.dept_id == dept_id).order_by(Semester.year, Semester.code).all()
        if not semesters:
            raise HTTPException(status_code=404, detail="No timetables found for department")
            
        excel_service = TimetableExcelService(db)
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        for sem in semesters:
            grid, allocations = excel_service._get_semester_allocations(sem.id)
            if not allocations: continue
            safe_title = sem.code[:31]
            suffix = 1
            original_title = safe_title
            while safe_title in wb.sheetnames:
                safe_title = f"{original_title[:28]}_{suffix}"
                suffix += 1
            ws = wb.create_sheet(title=safe_title)
            excel_service._build_semester_sheet(ws, sem, grid, allocations)
            
        if not wb.sheetnames:
            wb.create_sheet("Empty")
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=department_{dept_id}_timetables.xlsx"}
        )
    except Exception as e:
        logger.error(f"Department Excel export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# ELECTIVE AUDIT REPORT ENDPOINT
# ============================================================================

@router.get("/elective-audit")
def elective_audit_report(db: Session = Depends(get_db)):
    """
    Generate an elective audit report.
    Scans all allocations and reports:
    - Fake ELECTIVE entries
    - Elective group synchronization status
    - Missing class detection
    """
    from collections import defaultdict
    
    report = {
        "fake_elective_entries": [],
        "elective_groups": {},
        "missing_classes": [],
        "summary": {}
    }
    
    try:
        allocations = db.query(Allocation).options(
            joinedload(Allocation.subject),
            joinedload(Allocation.teacher),
            joinedload(Allocation.semester),
            joinedload(Allocation.room)
        ).filter(
            Allocation.is_elective == True,
            Allocation.subject_id.isnot(None)
        ).all()
        
        # Detect fake entries
        fake_entries = []
        for a in allocations:
            subj = a.subject
            if subj is None:
                fake_entries.append({
                    "allocation_id": a.id,
                    "semester_id": a.semester_id,
                    "day": a.day,
                    "slot": a.slot,
                    "reason": "subject_id is NULL"
                })
            elif subj.name in ["ELECTIVE", "OPEN ELECTIVE", "PROFESSIONAL ELECTIVE", "EMERGING ELECTIVE"]:
                fake_entries.append({
                    "allocation_id": a.id,
                    "semester_id": a.semester_id,
                    "day": a.day,
                    "slot": a.slot,
                    "subject_name": subj.name,
                    "subject_code": subj.code,
                    "teacher_name": _safe_name(a.teacher),
                    "reason": f"Fake subject name '{subj.name}'"
                })
        report["fake_elective_entries"] = fake_entries
        
        # Group by elective_basket_id
        groups = defaultdict(list)
        for a in allocations:
            if a.elective_basket_id is not None:
                groups[a.elective_basket_id].append(a)
        
        for basket_id, group_allocs in groups.items():
            # Check slot synchronization
            theory_slots = set()
            lab_slots = set()
            class_ids = set()
            
            for a in group_allocs:
                class_ids.add(a.semester_id)
                if a.component_type.value == "theory":
                    theory_slots.add((a.day, a.slot))
                elif a.component_type.value == "lab":
                    lab_slots.add((a.day, a.slot))
            
            report["elective_groups"][str(basket_id)] = {
                "basket_id": basket_id,
                "classes": list(class_ids),
                "theory_slots": [{"day": d, "slot": s} for d, s in sorted(theory_slots)],
                "lab_slots": [{"day": d, "slot": s} for d, s in sorted(lab_slots)],
                "theory_sync_ok": len(theory_slots) <= 1 if theory_slots else True,
                "lab_sync_ok": len(lab_slots) <= 1 if lab_slots else True,
                "total_allocations": len(group_allocs),
            }
        
        report["summary"] = {
            "total_elective_allocations": len(allocations),
            "fake_entries_count": len(fake_entries),
            "elective_groups_count": len(groups),
        }
        
    except Exception as e:
        report["error"] = str(e)
    
    return report


# ============================================================================
# DB CLEANUP ENDPOINT - REMOVE FAKE ELECTIVE RECORDS
# ============================================================================

@router.post("/cleanup-fake-electives")
def cleanup_fake_electives(db: Session = Depends(get_db)):
    """
    Remove fake ELECTIVE records from the database.
    Scans all allocations and removes entries where:
    - subject_name = "ELECTIVE" (or similar fake names)
    - subject_id is NULL
    - is_elective=True but no actual subject mapping
    
    Returns count of removed records.
    """
    try:
        fake_subject_names = ["ELECTIVE", "OPEN ELECTIVE", "PROFESSIONAL ELECTIVE", "EMERGING ELECTIVE"]
        
        # Find and delete allocations with fake subjects
        fake_allocs = db.query(Allocation).filter(
            Allocation.is_elective == True,
            Allocation.subject_id.isnot(None)
        ).join(Subject, Allocation.subject_id == Subject.id).filter(
            Subject.name.in_(fake_subject_names)
        ).all()
        
        fake_ids = [a.id for a in fake_allocs]
        removed_count = 0
        
        if fake_ids:
            removed_count = db.query(Allocation).filter(
                Allocation.id.in_(fake_ids)
            ).delete(synchronize_session='fetch')
            db.flush()
        
        # Also remove allocations with NULL subject_id that are marked as elective
        null_subject_allocs = db.query(Allocation).filter(
            Allocation.is_elective == True,
            Allocation.subject_id.is_(None)
        ).all()
        
        null_ids = [a.id for a in null_subject_allocs]
        if null_ids:
            removed_count += db.query(Allocation).filter(
                Allocation.id.in_(null_ids)
            ).delete(synchronize_session='fetch')
            db.commit()
        else:
            db.commit()
        
        return {
            "removed_count": removed_count,
            "message": f"Removed {removed_count} fake elective allocation(s)"
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Cleanup failed: {e}", exc_info=True)
        return {
            "removed_count": 0,
            "message": f"Cleanup failed: {str(e)}"
        }
# ============================================================================
# CLEAR ALLOCATIONS ENDPOINT
# ============================================================================

@router.delete('/clear', status_code=status.HTTP_204_NO_CONTENT)
def clear_allocations(
    semester_id: Optional[int] = None,
    dept_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Clear timetable allocations.
    If semester_id is provided, clears for that semester.
    If dept_id is provided, clears for that department.
    If neither is provided, clears all allocations.
    """
    try:
        query = db.query(Allocation)
        
        if semester_id:
            query = query.filter(Allocation.semester_id == semester_id)
        elif dept_id:
            query = query.filter(
                Allocation.semester_id.in_(
                    db.query(Semester.id).filter(Semester.department_id == dept_id)
                )
            )
            
        query.delete(synchronize_session=False)
        db.commit()
        
        # Invalidate caches
        cache.delete_pattern('timetable:*')
        cache.delete_pattern('allocations:*')
        
    except Exception as e:
        db.rollback()
        logger.error(f'Failed to clear allocations: {e}', exc_info=True)
        raise HTTPException(status_code=500, detail=f'Failed to clear allocations: {str(e)}')
