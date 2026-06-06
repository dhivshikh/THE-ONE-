"""
CRUD and import API routes for Parallel Lab Baskets.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session, joinedload

from app.db.session import get_db
from app.db.models import (
    Batch,
    Department,
    ParallelLabBasket,
    ParallelLabBasketSubject,
    Room,
    Semester,
    Subject,
    Teacher,
)

router = APIRouter(prefix="/parallel-lab-baskets", tags=["Parallel Lab Baskets"])


def _csv_ids(values: List[int]) -> str:
    return ",".join(str(v) for v in values if v)


def _parse_ids(value: Optional[str]) -> List[int]:
    if not value:
        return []
    ids = []
    for token in str(value).split(","):
        token = token.strip()
        if token.isdigit():
            ids.append(int(token))
    return ids


def _normalize_component(value: Optional[str]) -> str:
    normalized = (value or "lab").strip().lower().replace(" ", "_")
    if normalized in {"theory", "lab", "both"}:
        return normalized
    raise HTTPException(status_code=400, detail="Type must be Theory, Lab, or Both")


def _display_basket_name(basket: ParallelLabBasket) -> str:
    return basket.name or f"PB-{basket.dept_id}-Y{basket.year}-{basket.section}"


def _display_basket_code(basket: ParallelLabBasket) -> str:
    return basket.code or f"PLB{basket.id}"


def _subject_payload(row: ParallelLabBasketSubject) -> dict:
    lab_teacher_ids = _parse_ids(row.lab_teacher_ids)
    if not lab_teacher_ids and row.teacher_id:
        lab_teacher_ids = [row.teacher_id]

    return {
        "id": row.id,
        "basket_id": row.basket_id,
        "subject_id": row.subject_id,
        "batch_name": row.batch_name or "",
        "teacher_id": row.teacher_id,
        "theory_teacher_id": row.theory_teacher_id,
        "lab_teacher_ids": lab_teacher_ids,
        "room_id": row.room_id,
        "hours": row.hours or 2,
        "component_type": row.component_type or "lab",
        "subject": {
            "id": row.subject.id,
            "name": row.subject.name,
            "code": row.subject.code,
        } if row.subject else None,
        "teacher": {
            "id": row.teacher.id,
            "name": row.teacher.name,
        } if row.teacher else None,
        "theory_teacher": {
            "id": row.theory_teacher.id,
            "name": row.theory_teacher.name,
        } if getattr(row, "theory_teacher", None) else None,
        "lab_teachers": [],
        "room": {
            "id": row.room.id,
            "name": row.room.name,
        } if row.room else None,
    }


def _basket_payload(basket: ParallelLabBasket) -> dict:
    class_ids = _parse_ids(basket.class_ids)
    return {
        "id": basket.id,
        "name": _display_basket_name(basket),
        "code": _display_basket_code(basket),
        "dept_id": basket.dept_id,
        "year": basket.year,
        "section": basket.section,
        "semester_number": basket.semester_number,
        "class_ids": class_ids,
        "slot_day": basket.slot_day,
        "slot_period_start": basket.slot_period_start,
        "slot_period_count": basket.slot_period_count,
        "basket_subjects": [_subject_payload(s) for s in basket.basket_subjects],
    }


def _hydrate_lab_teacher_names(payload: dict, db: Session) -> dict:
    ids = {
        teacher_id
        for row in payload.get("basket_subjects", [])
        for teacher_id in row.get("lab_teacher_ids", [])
    }
    if not ids:
        return payload
    teachers = {
        t.id: {"id": t.id, "name": t.name}
        for t in db.query(Teacher).filter(Teacher.id.in_(ids)).all()
    }
    for row in payload.get("basket_subjects", []):
        row["lab_teachers"] = [teachers[tid] for tid in row.get("lab_teacher_ids", []) if tid in teachers]
    return payload


def _validate_subject_entry(entry: "ParallelLabBasketSubjectCreate", db: Session):
    component = _normalize_component(entry.component_type)
    if not db.query(Subject.id).filter(Subject.id == entry.subject_id).first():
        raise HTTPException(status_code=400, detail=f"Subject {entry.subject_id} not found")

    if component in {"theory", "both"}:
        if not entry.theory_teacher_id:
            raise HTTPException(status_code=400, detail="Theory Teacher Assigned validation failed")
        if not db.query(Teacher.id).filter(Teacher.id == entry.theory_teacher_id).first():
            raise HTTPException(status_code=400, detail=f"Theory teacher {entry.theory_teacher_id} not found")

    if component in {"lab", "both"}:
        lab_teacher_ids = entry.lab_teacher_ids or ([entry.teacher_id] if entry.teacher_id else [])
        if not lab_teacher_ids:
            raise HTTPException(status_code=400, detail="Lab Teachers Assigned validation failed")
        found = {
            t[0] for t in db.query(Teacher.id).filter(Teacher.id.in_(lab_teacher_ids)).all()
        }
        missing = [tid for tid in lab_teacher_ids if tid not in found]
        if missing:
            raise HTTPException(status_code=400, detail=f"Lab teacher(s) not found: {missing}")
        if not entry.room_id:
            raise HTTPException(status_code=400, detail="Room Assigned validation failed")
        if not db.query(Room.id).filter(Room.id == entry.room_id).first():
            raise HTTPException(status_code=400, detail=f"Room {entry.room_id} not found")


def _make_subject_row(basket_id: int, entry: "ParallelLabBasketSubjectCreate") -> ParallelLabBasketSubject:
    component = _normalize_component(entry.component_type)
    lab_teacher_ids = entry.lab_teacher_ids or ([entry.teacher_id] if entry.teacher_id else [])
    compatibility_teacher_id = None
    if lab_teacher_ids:
        compatibility_teacher_id = lab_teacher_ids[0]
    elif entry.theory_teacher_id:
        compatibility_teacher_id = entry.theory_teacher_id

    return ParallelLabBasketSubject(
        basket_id=basket_id,
        subject_id=entry.subject_id,
        batch_name=(entry.batch_name or "").strip(),
        teacher_id=compatibility_teacher_id,
        theory_teacher_id=entry.theory_teacher_id,
        lab_teacher_ids=_csv_ids(lab_teacher_ids),
        room_id=entry.room_id,
        hours=entry.hours,
        component_type=component,
    )


class ParallelLabBasketSubjectCreate(BaseModel):
    subject_id: int
    batch_name: str = ""
    teacher_id: Optional[int] = None
    theory_teacher_id: Optional[int] = None
    lab_teacher_ids: List[int] = Field(default_factory=list)
    room_id: Optional[int] = None
    hours: int = Field(default=2, ge=1, le=10)
    component_type: str = "lab"


class ParallelLabBasketCreate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    dept_id: int
    year: int
    section: str = ""
    semester_number: Optional[int] = None
    class_ids: List[int] = Field(default_factory=list)
    slot_day: int = -1
    slot_period_start: int = -1
    slot_period_count: int = 2
    subjects: List[ParallelLabBasketSubjectCreate]


class ParallelLabBasketSubjectResponse(BaseModel):
    id: int
    basket_id: int
    subject_id: int
    batch_name: str = ""
    teacher_id: Optional[int] = None
    theory_teacher_id: Optional[int] = None
    lab_teacher_ids: List[int] = []
    room_id: Optional[int] = None
    hours: int = 2
    component_type: str = "lab"
    subject: Optional[dict] = None
    teacher: Optional[dict] = None
    theory_teacher: Optional[dict] = None
    lab_teachers: List[dict] = []
    room: Optional[dict] = None


class ParallelLabBasketResponse(BaseModel):
    id: int
    name: str
    code: str
    dept_id: int
    year: int
    section: str
    semester_number: Optional[int] = None
    class_ids: List[int] = []
    slot_day: int
    slot_period_start: int
    slot_period_count: int
    basket_subjects: List[ParallelLabBasketSubjectResponse]


@router.get("/", response_model=List[ParallelLabBasketResponse])
def get_all_baskets(dept_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(ParallelLabBasket).options(
        joinedload(ParallelLabBasket.basket_subjects).joinedload(ParallelLabBasketSubject.subject),
        joinedload(ParallelLabBasket.basket_subjects).joinedload(ParallelLabBasketSubject.teacher),
        joinedload(ParallelLabBasket.basket_subjects).joinedload(ParallelLabBasketSubject.theory_teacher),
        joinedload(ParallelLabBasket.basket_subjects).joinedload(ParallelLabBasketSubject.room),
    )
    if dept_id:
        query = query.filter(ParallelLabBasket.dept_id == dept_id)
    return [_hydrate_lab_teacher_names(_basket_payload(b), db) for b in query.all()]


@router.post("/", response_model=ParallelLabBasketResponse)
def create_basket(basket_data: ParallelLabBasketCreate, db: Session = Depends(get_db)):
    if not basket_data.subjects:
        raise HTTPException(status_code=400, detail="Add at least one subject")
    for subj in basket_data.subjects:
        _validate_subject_entry(subj, db)

    if basket_data.class_ids:
        first_class = db.query(Semester).filter(Semester.id == basket_data.class_ids[0]).first()
        if not first_class:
            raise HTTPException(status_code=400, detail="Selected class not found")
        dept_id = first_class.dept_id or basket_data.dept_id
        year = first_class.year
        section = first_class.section
        semester_number = first_class.semester_number
    else:
        dept_id = basket_data.dept_id
        year = basket_data.year
        section = basket_data.section
        semester_number = basket_data.semester_number

    basket = ParallelLabBasket(
        name=(basket_data.name or "").strip() or None,
        code=(basket_data.code or "").strip() or None,
        dept_id=dept_id,
        year=year,
        section=section,
        semester_number=semester_number,
        class_ids=_csv_ids(basket_data.class_ids),
        slot_day=basket_data.slot_day,
        slot_period_start=basket_data.slot_period_start,
        slot_period_count=basket_data.slot_period_count,
    )
    db.add(basket)
    db.flush()

    for subj in basket_data.subjects:
        db.add(_make_subject_row(basket.id, subj))

    db.commit()
    db.refresh(basket)
    return _hydrate_lab_teacher_names(_basket_payload(basket), db)


@router.delete("/{basket_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_basket(basket_id: int, db: Session = Depends(get_db)):
    basket = db.query(ParallelLabBasket).filter(ParallelLabBasket.id == basket_id).first()
    if not basket:
        raise HTTPException(status_code=404, detail="Basket not found")

    db.delete(basket)
    db.commit()
    return None


@router.put("/{basket_id}", response_model=ParallelLabBasketResponse)
def update_basket(basket_id: int, basket_data: ParallelLabBasketCreate, db: Session = Depends(get_db)):
    basket = db.query(ParallelLabBasket).filter(ParallelLabBasket.id == basket_id).first()
    if not basket:
        raise HTTPException(status_code=404, detail="Basket not found")
    if not basket_data.subjects:
        raise HTTPException(status_code=400, detail="Add at least one subject")
    for subj in basket_data.subjects:
        _validate_subject_entry(subj, db)

    basket.name = (basket_data.name or "").strip() or basket.name
    basket.code = (basket_data.code or "").strip() or basket.code
    basket.dept_id = basket_data.dept_id
    basket.year = basket_data.year
    basket.section = basket_data.section
    basket.semester_number = basket_data.semester_number
    basket.class_ids = _csv_ids(basket_data.class_ids)
    basket.slot_day = basket_data.slot_day
    basket.slot_period_start = basket_data.slot_period_start
    basket.slot_period_count = basket_data.slot_period_count

    db.query(ParallelLabBasketSubject).filter(ParallelLabBasketSubject.basket_id == basket_id).delete()
    for subj in basket_data.subjects:
        db.add(_make_subject_row(basket.id, subj))

    db.commit()
    db.refresh(basket)
    return _hydrate_lab_teacher_names(_basket_payload(basket), db)


# ============================================================================
# Import support
# ============================================================================

IMPORT_COLUMNS = ["Basket Name", "Class", "Subject", "Type", "Teacher", "Room"]
OPTIONAL_IMPORT_COLUMNS = ["Basket Code", "Batch"]
_pending_imports: Dict[str, dict] = {}


def _norm(value: str) -> str:
    return " ".join((value or "").strip().upper().split())


def _parse_import_file(file_bytes: bytes, filename: str) -> List[dict]:
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        try:
            import openpyxl
        except ImportError as exc:
            raise HTTPException(status_code=400, detail="openpyxl is not installed") from exc
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(v).strip() if v is not None else "" for v in rows[0]]
        data_rows = rows[1:]
    elif lower.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            return []
        headers = [h.strip() for h in rows[0]]
        data_rows = rows[1:]
    else:
        raise HTTPException(status_code=400, detail="Use .xlsx, .xls, or .csv")

    header_map = {_norm(h): i for i, h in enumerate(headers)}
    missing = [col for col in IMPORT_COLUMNS if _norm(col) not in header_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")

    parsed = []
    for row_num, row in enumerate(data_rows, start=2):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        item = {"row": row_num}
        for col in IMPORT_COLUMNS + OPTIONAL_IMPORT_COLUMNS:
            idx = header_map.get(_norm(col), -1)
            value = row[idx] if idx >= 0 and idx < len(row) else ""
            item[col] = str(value).strip() if value is not None else ""
        parsed.append(item)
    return parsed


def _resolve_import_rows(rows: List[dict], db: Session) -> List[dict]:
    classes = {}
    for sem in db.query(Semester).all():
        classes[_norm(sem.code)] = sem
        classes[_norm(sem.name)] = sem
    subjects = {}
    for subj in db.query(Subject).all():
        subjects[_norm(subj.code)] = subj
        subjects[_norm(subj.name)] = subj
    teachers = {}
    for teacher in db.query(Teacher).all():
        if teacher.teacher_code:
            teachers[_norm(teacher.teacher_code)] = teacher
        teachers[_norm(teacher.name)] = teacher
    rooms = {_norm(room.name): room for room in db.query(Room).all()}

    resolved = []
    for row in rows:
        errors = []
        sem = classes.get(_norm(row["Class"]))
        subj = subjects.get(_norm(row["Subject"]))
        teacher = teachers.get(_norm(row["Teacher"]))
        room = rooms.get(_norm(row["Room"])) if row.get("Room") else None
        comp = (row["Type"] or "").strip().lower()
        if comp not in {"theory", "lab"}:
            errors.append("Type must be Theory or Lab")
        if not row["Basket Name"]:
            errors.append("Basket Name is required")
        if not sem:
            errors.append(f"Class '{row['Class']}' not found")
        if not subj:
            errors.append(f"Subject '{row['Subject']}' not found")
        if not teacher:
            errors.append(f"Teacher '{row['Teacher']}' not found")
        if comp == "lab" and not room:
            errors.append(f"Lab room '{row['Room']}' not found")

        resolved.append({
            "row": row["row"],
            "data": row,
            "errors": errors,
            "status": "invalid" if errors else "valid",
            "semester_id": sem.id if sem else None,
            "subject_id": subj.id if subj else None,
            "teacher_id": teacher.id if teacher else None,
            "room_id": room.id if room else None,
        })
    return resolved


@router.post("/import/upload")
async def upload_import(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    contents = await file.read()
    rows = _parse_import_file(contents, file.filename)
    resolved = _resolve_import_rows(rows, db)
    batch_id = f"parallel_basket_import_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    _pending_imports[batch_id] = {"file_bytes": contents, "filename": file.filename}
    failed = sum(1 for r in resolved if r["status"] == "invalid")
    return {
        "batch_id": batch_id,
        "total_rows": len(resolved),
        "failed": failed,
        "rows": resolved,
        "schema_errors": [],
    }


@router.post("/import/commit")
async def commit_import(batch_id: str, db: Session = Depends(get_db)):
    pending = _pending_imports.get(batch_id)
    if not pending:
        raise HTTPException(status_code=404, detail="Import batch not found")

    rows = _parse_import_file(pending["file_bytes"], pending["filename"])
    resolved = _resolve_import_rows(rows, db)
    valid_rows = [r for r in resolved if r["status"] == "valid"]

    created_baskets = 0
    created_entries = 0
    updated_entries = 0

    for row in valid_rows:
        data = row["data"]
        sem = db.query(Semester).filter(Semester.id == row["semester_id"]).first()
        if not sem:
            continue
        basket_name = data["Basket Name"].strip()
        basket_code = (data.get("Basket Code") or basket_name).strip()
        basket = db.query(ParallelLabBasket).filter(
            ParallelLabBasket.name == basket_name,
            ParallelLabBasket.dept_id == sem.dept_id,
        ).first()
        if not basket:
            basket = ParallelLabBasket(
                name=basket_name,
                code=basket_code,
                dept_id=sem.dept_id,
                year=sem.year,
                section=sem.section,
                semester_number=sem.semester_number,
                class_ids=str(sem.id),
                slot_day=-1,
                slot_period_start=-1,
                slot_period_count=2,
            )
            db.add(basket)
            db.flush()
            created_baskets += 1
        else:
            class_ids = set(_parse_ids(basket.class_ids))
            if sem.id not in class_ids:
                class_ids.add(sem.id)
                basket.class_ids = _csv_ids(sorted(class_ids))

        batch_name = (data.get("Batch") or "").strip()
        entry = db.query(ParallelLabBasketSubject).filter(
            ParallelLabBasketSubject.basket_id == basket.id,
            ParallelLabBasketSubject.subject_id == row["subject_id"],
            ParallelLabBasketSubject.batch_name == batch_name,
        ).first()
        if not entry:
            entry = ParallelLabBasketSubject(
                basket_id=basket.id,
                subject_id=row["subject_id"],
                batch_name=batch_name,
                teacher_id=row["teacher_id"],
                component_type=data["Type"].strip().lower(),
                hours=2,
            )
            db.add(entry)
            created_entries += 1
        else:
            updated_entries += 1

        if data["Type"].strip().lower() == "theory":
            entry.theory_teacher_id = row["teacher_id"]
            if entry.component_type == "lab":
                entry.component_type = "both"
            else:
                entry.component_type = "theory"
            if not entry.teacher_id:
                entry.teacher_id = row["teacher_id"]
        else:
            existing_lab_ids = set(_parse_ids(entry.lab_teacher_ids))
            existing_lab_ids.add(row["teacher_id"])
            entry.lab_teacher_ids = _csv_ids(sorted(existing_lab_ids))
            entry.teacher_id = entry.teacher_id or row["teacher_id"]
            entry.room_id = row["room_id"]
            if entry.component_type == "theory":
                entry.component_type = "both"
            else:
                entry.component_type = "lab"

    db.commit()
    _pending_imports.pop(batch_id, None)
    return {
        "total_rows": len(resolved),
        "imported_rows": len(valid_rows),
        "failed": len(resolved) - len(valid_rows),
        "created_baskets": created_baskets,
        "created_entries": created_entries,
        "updated_entries": updated_entries,
        "rows": resolved,
    }


@router.get("/import/template")
async def download_import_template():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is not installed") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PARALLEL_BASKETS"
    columns = ["Basket Name", "Basket Code", "Class", "Subject", "Type", "Teacher", "Room", "Batch"]
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
        ws.column_dimensions[cell.column_letter].width = 18
    examples = [
        ["AIDS_Y2_PB1", "AIDS_Y2_PB1", "DS2A", "DAA", "Theory", "Teacher Z", "", ""],
        ["AIDS_Y2_PB1", "AIDS_Y2_PB1", "DS2A", "DAA", "Lab", "Teacher X", "AI Lab 1", "B1"],
        ["AIDS_Y2_PB1", "AIDS_Y2_PB1", "DS2A", "DAA", "Lab", "Teacher Y", "AI Lab 1", "B2"],
    ]
    for r_idx, row in enumerate(examples, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=parallel_basket_import_template.xlsx"},
    )
